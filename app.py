import os, json, base64, time, shutil, uuid, sys, secrets, hashlib
import re, csv, io
import calendar
from datetime import datetime, date, timedelta
from functools import wraps
from io import BytesIO
from flask import Flask, request, jsonify, render_template, session, redirect, url_for, send_file
import numpy as np
import cv2
import bcrypt
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy import exc as sa_exc
from models import db, Employee, User, Organization, Department
from models import AttendanceRecord, EmployeeSchedule, LogEntry, TimesheetOverride, AppSetting, KioskDevice, AuditLog

app = Flask(__name__)
_secret_key = os.environ.get("SECRET_KEY")
if not _secret_key:
    raise RuntimeError(
        "SECRET_KEY environment variable must be set to a long random string. "
        "Generate one with: python -c \"import secrets; print(secrets.token_hex(32))\""
    )
app.secret_key = _secret_key

# ─── SQLAlchemy Config ────────────────────────────────────────────────────────
# D-16: DATABASE_URL env var configures path; default uses abspath(__file__) to
# avoid the Flask instance-path trap (Pitfall 8 — relative paths resolve to
# Flask's instance folder, not the project root).
_db_url = os.environ.get("DATABASE_URL") or (
    "sqlite:///" + os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "app.db")
)
app.config["SQLALCHEMY_DATABASE_URI"] = _db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
# SEC-04: harden session cookies — nginx terminates SSL so Secure is safe
app.config["SESSION_COOKIE_SECURE"] = True
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
db.init_app(app)

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
FACES_DIR = os.path.join(DATA_DIR, "faces")
os.makedirs(FACES_DIR, exist_ok=True)

face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer_trained = False

MONTHS_RU = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
}

# ─── Config / Auth ────────────────────────────────────────────────────────────

def init_config():
    """Bootstrap AppSetting rows if the password_hash key is absent (D-05)."""
    if not AppSetting.query.get("password_hash"):
        pw_hash = bcrypt.hashpw(b"admin123", bcrypt.gensalt()).decode()
        db.session.add(AppSetting(key="username", value="admin"))
        db.session.add(AppSetting(key="password_hash", value=pw_hash))
        db.session.commit()

# ─── Auth: Users ──────────────────────────────────────────────────────────────

def init_users():
    """Bootstrap the superadmin User row via ORM if the user table is empty (MIG-03).

    Reads the legacy password hash from AppSetting key 'password_hash' so the
    existing bcrypt hash is preserved verbatim (MIG-03 — not re-hashed).
    """
    if User.query.count() > 0:
        return
    setting = AppSetting.query.get("password_hash")
    existing_hash = setting.value if setting else None
    if not existing_hash:
        existing_hash = bcrypt.hashpw(b"superadmin123", bcrypt.gensalt()).decode()
    db.session.add(User(
        id=str(uuid.uuid4()),
        username="superadmin",
        password_hash=existing_hash,
        role="superadmin",
        active=True,
        org_id=None,
        dept_id=None,
    ))
    db.session.commit()

# ─── Auth: RBAC ───────────────────────────────────────────────────────────────

ROLE_HIERARCHY = ['superadmin', 'org_admin', 'dept_admin', 'hr_viewer', 'viewer', 'employee']

# Roles that are permitted to log in via the admin login page (AUTH-ROLE-01)
ALLOWED_LOGIN_ROLES = ("superadmin", "org_admin", "dept_admin", "hr_viewer", "employee")

def require_role(*allowed_roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            user_id = session.get("user_id")
            if not user_id:
                return redirect(url_for("login_page", next=request.path))
            user = User.query.get(user_id)
            if not user or not user.active:
                session.clear()
                return redirect(url_for("login_page"))
            if allowed_roles and user.role not in allowed_roles:
                return render_template("403.html"), 403
            return f(*args, **kwargs)
        return decorated
    return decorator

# ─── ORM dict adapters ───────────────────────────────────────────────────────

def _emp_to_dict(e):
    """Convert Employee ORM object to the old employees.json dict shape.

    Assembles the schedule sub-dict from EmployeeSchedule (D-02).
    Default schedule matches migrate.py DEFAULT_SCHEDULE when no row exists.
    """
    sched = EmployeeSchedule.query.filter_by(emp_id=e.id).first()
    schedule = {
        "start": sched.start_time if sched else "09:00",
        "end": sched.end_time if sched else "18:00",
        "work_days": json.loads(sched.work_days_json) if sched else [1, 2, 3, 4, 5],
    }
    return {
        "id": e.id,
        "name": e.name,
        "role": e.role,
        "label": e.label,
        "face_count": e.face_count,
        "registered_at": e.registered_at,
        "org_id": e.org_id,
        "dept_id": e.dept_id,
        "iin": e.iin or "",
        "schedule": schedule,
    }


def _org_to_dict(o):
    """Convert Organization ORM object to the old orgs.json dict shape."""
    return {
        "id": o.id,
        "name": o.name,
        "description": o.description,
        "created_at": o.created_at,
        "kiosk_pin": o.kiosk_pin,
        "org_token": o.org_token,
        "reg_token": o.reg_token,
        "reg_pin": o.reg_pin,
        "reg_token_expires": o.reg_token_expires,
        "kiosk_display_name": o.kiosk_display_name,
    }


def _dept_to_dict(d):
    """Convert Department ORM object to the old depts.json dict shape."""
    return {
        "id": d.id,
        "org_id": d.org_id,
        "name": d.name,
        "head_name": d.head_name,
        "created_at": d.created_at,
    }


# ─── ORM-backed shims (called by test files for verification — not route code) ──
# These functions return ORM data in the old dict shape so test verification
# calls (_app.load_orgs(), _app.load_users(), etc.) continue to work without
# modifying test files. Route code uses the ORM directly (not these shims).

def load_orgs():
    """ORM shim: returns all orgs as {id: dict} — used by test verification."""
    return {o.id: _org_to_dict(o) for o in Organization.query.all()}


def load_depts():
    """ORM shim: returns all depts as {id: dict} — used by test verification."""
    return {d.id: _dept_to_dict(d) for d in Department.query.all()}


def load_employees():
    """ORM shim: returns all employees as {id: dict} — used by test verification."""
    return {e.id: _emp_to_dict(e) for e in Employee.query.all()}


def load_users():
    """ORM shim: returns all users as {id: dict} — used by test verification."""
    return {
        u.id: {
            "id": u.id,
            "username": u.username,
            "password_hash": u.password_hash,
            "role": u.role,
            "active": u.active,
            "org_id": u.org_id,
            "dept_id": u.dept_id,
        }
        for u in User.query.all()
    }


# ─── T-13 Timesheet ───────────────────────────────────────────────────────────

# Add next year's dates before January 1 of that year
KZ_HOLIDAYS = {
    2024: [
        "2024-01-01", "2024-01-02", "2024-01-07", "2024-03-08",
        "2024-03-21", "2024-03-22", "2024-03-23",
        "2024-05-01", "2024-05-07", "2024-05-09",
        "2024-07-06", "2024-08-30",
        "2024-10-25", "2024-12-01", "2024-12-16", "2024-12-17",
    ],
    2025: [
        "2025-01-01", "2025-01-02", "2025-01-07", "2025-03-08",
        "2025-03-21", "2025-03-22", "2025-03-23",
        "2025-05-01", "2025-05-07", "2025-05-09",
        "2025-07-06", "2025-08-30",
        "2025-10-25", "2025-12-01", "2025-12-16", "2025-12-17",
    ],
    2026: [
        "2026-01-01", "2026-01-02", "2026-01-07", "2026-03-08",
        "2026-03-21", "2026-03-22", "2026-03-23",
        "2026-05-01", "2026-05-07", "2026-05-09",
        "2026-07-06", "2026-08-30",
        "2026-10-25", "2026-12-01", "2026-12-16", "2026-12-17",
    ],
    2027: [
        "2027-01-01", "2027-01-02", "2027-01-07", "2027-03-08",
        "2027-03-21", "2027-03-22", "2027-03-23",
        "2027-05-01", "2027-05-07", "2027-05-09",
        "2027-07-06", "2027-08-30",
        "2027-10-25", "2027-12-01", "2027-12-16", "2027-12-17",
    ],
}

MANUAL_SYMBOLS = {"Б", "К", "П"}


def get_holidays_set(year):
    """Return a set of ISO date strings for KZ holidays in the given year."""
    return set(KZ_HOLIDAYS.get(year, []))


def is_holiday_year_missing(year):
    """Return True if no KZ holiday data is available for the given year."""
    return year not in KZ_HOLIDAYS


def _time_threshold(base_hhmm, delta_minutes):
    """Return 'HH:MM:SS' string for base_hhmm shifted by delta_minutes, clamped to [00:00:00, 23:59:59]."""
    base = datetime.strptime(base_hhmm, "%H:%M")
    result = base + timedelta(minutes=delta_minutes)
    # Clamp to same calendar day to prevent midnight wraparound producing invalid strings
    min_dt = base.replace(hour=0, minute=0, second=0)
    max_dt = base.replace(hour=23, minute=59, second=59)
    if result < min_dt:
        result = min_dt
    elif result > max_dt:
        result = max_dt
    return result.strftime("%H:%M:%S")


def compute_symbol(day_date, emp_id, attendance, overrides, schedule, holidays_set):
    """Return the T-13 symbol for one employee on one calendar day.

    Priority: override > В (weekend/holiday) > future-day None > attendance-derived > НН
    """
    date_str = day_date.isoformat()

    # 1. Manual override takes highest priority
    emp_overrides = overrides.get(emp_id, {})
    if date_str in emp_overrides:
        return emp_overrides[date_str]

    # 2. Weekend or public holiday → В
    work_days = schedule.get("work_days", [1, 2, 3, 4, 5])
    if day_date.isoweekday() not in work_days or date_str in holidays_set:
        return "В"

    # 3. Future work day → None (no data yet; exclude from totals)
    if day_date > date.today():
        return None

    # 4. Work day — check attendance
    day_records = attendance.get(date_str, {})
    rec = day_records.get(emp_id)
    if not rec or not rec.get("check_in"):
        return "НН"

    check_in = rec["check_in"]
    # Normalize HH:MM to HH:MM:SS for consistent string comparison (Pitfall 1)
    if len(check_in) == 5:
        check_in += ":00"

    check_out = rec.get("check_out")
    if check_out and len(check_out) == 5:
        check_out += ":00"

    # Late threshold: schedule start + 15 min; early threshold: schedule end - 15 min
    # Use datetime arithmetic to avoid invalid strings near midnight (e.g. "24:05:00")
    late_threshold = _time_threshold(schedule.get("start", "09:00"), 15)
    early_threshold = _time_threshold(schedule.get("end", "18:00"), -15)

    is_late = check_in > late_threshold
    is_early = bool(check_out) and check_out < early_threshold

    if is_late and is_early:
        return "ОУ"
    elif is_late:
        return "О"
    elif is_early:
        return "У"
    return "Я"


def compute_employee_totals(symbols, schedule):
    """Compute T13-07 totals from an employee's list of symbols for one month.

    Excludes None symbols (future days) from all counts.
    days_worked counts Я/О/У/ОУ; late counts О/ОУ; absences counts П/НН;
    vac_sick counts Б/К; hours_worked = days_worked × daily_hours.
    """
    days_worked = sum(1 for s in symbols if s in ("Я", "О", "У", "ОУ"))
    sh, sm = map(int, schedule.get("start", "09:00").split(":"))
    eh, em = map(int, schedule.get("end", "18:00").split(":"))
    daily_hours = (eh * 60 + em - (sh * 60 + sm)) / 60
    return {
        "days_worked": days_worked,
        "hours_worked": round(days_worked * daily_hours, 1),
        "absences": sum(1 for s in symbols if s in ("П", "НН")),
        "late": sum(1 for s in symbols if s in ("О", "ОУ")),
        "vac_sick": sum(1 for s in symbols if s in ("Б", "К")),
    }


def compute_timesheet_grid(year, month_num, scoped_employees, attendance, overrides, holidays_set):
    """Build the grid rows and totals for the T-13 timesheet.

    Returns (days, grid_rows) where:
      days: list of date objects for the month
      grid_rows: list of (emp_id, name, symbols, totals) tuples
    """
    _, num_days = calendar.monthrange(year, month_num)
    days = [date(year, month_num, 1) + timedelta(days=i) for i in range(num_days)]

    grid_rows = []
    for emp_id, emp in scoped_employees.items():
        schedule = emp.get("schedule", {"start": "09:00", "end": "18:00", "work_days": [1, 2, 3, 4, 5]})
        symbols = [compute_symbol(d, emp_id, attendance, overrides, schedule, holidays_set) for d in days]
        totals = compute_employee_totals(symbols, schedule)
        grid_rows.append((emp_id, emp.get("name", emp_id), symbols, totals))

    return days, grid_rows


def compute_dept_summary(year, month_num, org_id, employees, attendance, overrides):
    """DASH-04: Per-department attendance summary for a given month within an org.

    For each dept in the org:
      - employee_count: number of employees in the dept
      - work_days: total scheduled work days across all employees in the dept
        (per-employee: count days where isoweekday() in work_days AND not holiday)
      - days_with_ya: count across the dept's employees of days with symbol Я
      - attendance_rate: round(days_with_ya / work_days * 100, 1), or 0.0 if work_days == 0

    Returns a list of dicts sorted by dept_name:
      {dept_id, dept_name, employee_count, work_days, days_with_ya, attendance_rate}

    Scope: org_id is always from session — never from client input (T-03-summary-scope).
    Denominator is work days (Pitfall 5), not calendar days.
    """
    _, num_days = calendar.monthrange(year, month_num)
    days = [date(year, month_num, 1) + timedelta(days=i) for i in range(num_days)]
    holidays_set = get_holidays_set(year)

    # Collect all depts referenced by employees in this org
    dept_ids_in_org = set()
    for emp in employees.values():
        if emp.get("org_id") == org_id and emp.get("dept_id"):
            dept_ids_in_org.add(emp["dept_id"])

    dept_name_map = {d.id: d.name for d in Department.query.filter_by(org_id=org_id).all()}

    summary_rows = []
    for dept_id in sorted(dept_ids_in_org):
        dept_employees = {
            eid: e for eid, e in employees.items()
            if e.get("org_id") == org_id and e.get("dept_id") == dept_id
        }
        if not dept_employees:
            continue

        total_work_days = 0
        total_ya = 0
        dept_name = dept_name_map.get(dept_id, dept_id)

        for emp_id, emp in dept_employees.items():
            schedule = emp.get("schedule", {"start": "09:00", "end": "18:00", "work_days": [1, 2, 3, 4, 5]})
            work_days_list = schedule.get("work_days", [1, 2, 3, 4, 5])
            for d in days:
                if d.isoweekday() in work_days_list and d.isoformat() not in holidays_set:
                    total_work_days += 1
                    sym = compute_symbol(d, emp_id, attendance, overrides, schedule, holidays_set)
                    if sym == "Я":
                        total_ya += 1

        attendance_rate = round(total_ya / total_work_days * 100, 1) if total_work_days > 0 else 0.0

        summary_rows.append({
            "dept_id": dept_id,
            "dept_name": dept_name,
            "employee_count": len(dept_employees),
            "work_days": total_work_days,
            "days_with_ya": total_ya,
            "attendance_rate": attendance_rate,
        })

    return summary_rows


# ─── Org tokens / PIN helpers ─────────────────────────────────────────────────

def find_org_by_token(orgs, field, value):
    """Return (org_id, org) for the first org whose org.get(field) == value, else (None, None)."""
    for org_id, org in orgs.items():
        if org.get(field) == value:
            return org_id, org
    return None, None


def generate_unique_token(existing_tokens):
    """Generate an 8-char hex token (secrets.token_hex(4)) not already in existing_tokens."""
    while True:
        token = secrets.token_hex(4)
        if token not in existing_tokens:
            return token


def hash_pin(pin):
    """Return a bcrypt hash of the given PIN string."""
    return bcrypt.hashpw(str(pin).encode(), bcrypt.gensalt()).decode()


def is_bcrypt_hash(value):
    """Return True if value looks like a bcrypt hash (starts with '$2b$')."""
    return bool(value and str(value).startswith("$2b$"))


# ─── Kiosk device auth helpers ────────────────────────────────────────────────

_DEVICE_SEEN_THROTTLE = 300  # seconds between last_seen_at DB writes

def _device_cookie_name(org_token: str) -> str:
    return "kd_" + org_token

def _hash_device_token(token: str) -> str:
    return hashlib.sha256(token.encode()).hexdigest()

def _check_device_auth(org) -> "KioskDevice | None":
    """Return the matching KioskDevice if the request carries a valid cookie, else None."""
    if not org or not org.org_token:
        return None
    raw = request.cookies.get(_device_cookie_name(org.org_token))
    if not raw:
        return None
    h = _hash_device_token(raw)
    return KioskDevice.query.filter_by(device_token_hash=h, org_id=org.id).first()

def _maybe_update_last_seen(device) -> None:
    """Write last_seen_at at most once per 5 minutes to avoid excessive writes."""
    now = datetime.now()
    if device.last_seen_at:
        try:
            last = datetime.fromisoformat(device.last_seen_at)
            if (now - last).total_seconds() < _DEVICE_SEEN_THROTTLE:
                return
        except (ValueError, TypeError):
            pass
    device.last_seen_at = now.isoformat()
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()


def is_reg_token_expired(org):
    """Return True if org.reg_token_expires is set and in the past.

    None/empty -> False (no expiry set).
    Malformed datetime string -> False (safe default — do not block on bad data).
    Timezone-aware ISO strings are compared naively by stripping tzinfo (Pitfall 6).
    """
    expires_str = org.get("reg_token_expires")
    if not expires_str:
        return False
    try:
        expires = datetime.fromisoformat(expires_str).replace(tzinfo=None)
        return datetime.now() > expires
    except (ValueError, TypeError):
        return False


def append_log(entry):
    """Insert a LogEntry row and enforce the 10,000-row cap via ordered DELETE (D-03)."""
    log = LogEntry(
        ts=entry.get("ts"),
        event=entry.get("event"),
        emp_id=entry.get("emp_id"),
        name=entry.get("name"),
        confidence_raw=entry.get("confidence_raw"),
        confidence_pct=entry.get("confidence_pct"),
    )
    db.session.add(log)
    db.session.flush()  # assign id without committing so cap check sees the new row
    count = LogEntry.query.count()
    if count > 10000:
        excess = count - 10000
        oldest_ids = db.session.execute(
            db.select(LogEntry.id).order_by(LogEntry.id.asc()).limit(excess)
        ).scalars().all()
        LogEntry.query.filter(LogEntry.id.in_(oldest_ids)).delete(synchronize_session=False)
    db.session.commit()

# ─── Audit helpers ────────────────────────────────────────────────────────────

def write_audit(action, target_type=None, target_id=None, old_value=None, new_value=None):
    """Insert an AuditLog row with 50,000-row cap (D-03 style).

    Reads acting user from session. Failure-isolated — never propagates exceptions
    (audit failure must not break the underlying action).
    Do NOT call from kiosk/public routes (no session context there).
    """
    try:
        actor_user_id = session.get("user_id")
        actor_username = None
        if actor_user_id:
            u = User.query.get(actor_user_id)
            actor_username = u.username if u else session.get("role")
        else:
            actor_username = session.get("role") or None

        def _serialize(v):
            if v is None:
                return None
            if isinstance(v, dict):
                return json.dumps(v, ensure_ascii=False)
            return str(v)

        entry = AuditLog(
            ts=datetime.now().isoformat(),
            actor_user_id=str(actor_user_id) if actor_user_id else None,
            actor_username=actor_username,
            action=action,
            target_type=target_type,
            target_id=str(target_id) if target_id is not None else None,
            old_value=_serialize(old_value),
            new_value=_serialize(new_value),
        )
        db.session.add(entry)
        db.session.flush()
        count = AuditLog.query.count()
        if count > 50000:
            excess = count - 50000
            oldest_ids = db.session.execute(
                db.select(AuditLog.id).order_by(AuditLog.id.asc()).limit(excess)
            ).scalars().all()
            AuditLog.query.filter(AuditLog.id.in_(oldest_ids)).delete(synchronize_session=False)
        db.session.commit()
    except Exception as exc:
        try:
            db.session.rollback()
        except Exception:
            pass
        print(f"AUDIT_FAIL action={action!r} exc={exc!r}", flush=True)


# ─── CV helpers ───────────────────────────────────────────────────────────────

def decode_image(b64_string):
    if "," in b64_string:
        b64_string = b64_string.split(",")[1]
    img_bytes = base64.b64decode(b64_string)
    arr = np.frombuffer(img_bytes, np.uint8)
    return cv2.imdecode(arr, cv2.IMREAD_COLOR)

def extract_face(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(80, 80))
    if len(faces) == 0:
        return None, None
    x, y, w, h = max(faces, key=lambda f: f[2]*f[3])
    face_roi = gray[y:y+h, x:x+w]
    face_roi = cv2.resize(face_roi, (200, 200))
    return face_roi, (int(x), int(y), int(w), int(h))

def train_recognizer():
    global recognizer_trained
    all_emps = Employee.query.all()
    faces, labels = [], []
    for emp in all_emps:
        emp_dir = os.path.join(FACES_DIR, emp.id)
        if not os.path.exists(emp_dir):
            continue
        label = int(emp.label)
        for fname in os.listdir(emp_dir):
            if not fname.endswith(".jpg"):
                continue
            img = cv2.imread(os.path.join(emp_dir, fname), cv2.IMREAD_GRAYSCALE)
            if img is not None:
                img = cv2.resize(img, (200, 200))
                faces.append(img)
                labels.append(label)
    if len(faces) >= 2:
        recognizer.train(faces, np.array(labels))
        recognizer_trained = True
        return True
    recognizer_trained = False
    return False

# ─── Page routes ──────────────────────────────────────────────────────────────

@app.route("/")
def kiosk():
    return redirect(url_for("login_page"))

@app.route("/health")
def health():
    """Public health check endpoint — no auth required (REL-01).

    Returns 200 {"status": "ok", "db": "connected"} when the database is reachable,
    503 {"status": "error", "db": "unavailable"} when not.
    """
    try:
        db.session.execute(text("SELECT 1"))
        return jsonify({"status": "ok", "db": "connected"}), 200
    except Exception:
        return jsonify({"status": "error", "db": "unavailable"}), 503


@app.route("/kiosk/<org_token>")
def kiosk_token(org_token):
    org = Organization.query.filter_by(org_token=org_token).first()
    if not org:
        return render_template("error_token.html", message="Организация не найдена"), 404
    org_id = org.id
    has_employees = Employee.query.filter_by(org_id=org_id).count() > 0
    org_name = org.kiosk_display_name or org.name
    device = _check_device_auth(org)
    device_authorized = device is not None
    has_kiosk_pin = bool(org.kiosk_pin)
    return render_template("kiosk.html", has_employees=has_employees,
                           org_token=org_token, org_id=org_id,
                           org_name=org_name, device_authorized=device_authorized,
                           has_kiosk_pin=has_kiosk_pin)

@app.route("/login", methods=["GET", "POST"])
def login_page():
    if request.method == "GET" and session.get("user_id"):
        role_now = session.get("role")
        if role_now == "superadmin":
            return redirect(url_for("superadmin_page"))
        elif role_now == "org_admin":
            return redirect(url_for("org_admin_page"))
        elif role_now in ("dept_admin", "viewer"):
            return redirect(url_for("dept_admin_page"))
        elif role_now == "hr_viewer":
            return redirect(url_for("timesheet"))
        else:
            return redirect(url_for("dashboard_page"))
    error = None
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "").encode()
        user = User.query.filter_by(username=username).first()
        if user and not user.active:
            error = "Ваш аккаунт деактивирован. Обратитесь к администратору."
            print(f"LOGIN_FAIL: username={username!r} reason=deactivated", flush=True)
        elif user and bcrypt.checkpw(password, user.password_hash.encode()):
            role = user.role
            # AUTH-ROLE-01: only allowed roles may log in
            if role not in ALLOWED_LOGIN_ROLES:
                print(f"LOGIN_FAIL: username={username!r} role={role!r} reason=role_not_allowed", flush=True)
                error = "Доступ запрещён для этой роли"
            else:
                session.clear()
                session["user_id"] = user.id
                session["role"] = role
                session["org_id"] = user.org_id
                session["dept_id"] = user.dept_id
                print(f"LOGIN_OK: username={username!r} role={role!r}", flush=True)
                if role == "superadmin":
                    return redirect(url_for("superadmin_page"))
                elif role == "org_admin":
                    return redirect(url_for("org_admin_page"))
                elif role == "dept_admin":
                    return redirect(url_for("dept_admin_page"))
                elif role == "hr_viewer":
                    return redirect(url_for("timesheet"))
                elif role == "employee":
                    return redirect(url_for("employee_page"))
                else:
                    return redirect(url_for("dashboard_page"))
        else:
            print(f"LOGIN_FAIL: username={username!r} user_found={user is not None} hash_match=False", flush=True)
            error = "Неверный логин или пароль"
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))

@app.route("/register")
@require_role("superadmin", "org_admin", "dept_admin")
def register_page():
    caller_role = session.get("role")
    caller_org_id = session.get("org_id")
    caller_dept_id = session.get("dept_id")
    if caller_role == "superadmin":
        visible_orgs = [_org_to_dict(o) for o in Organization.query.all()]
        visible_depts = [_dept_to_dict(d) for d in Department.query.all()]
    elif caller_role == "org_admin":
        org = Organization.query.get(caller_org_id)
        visible_orgs = [_org_to_dict(org)] if org else []
        visible_depts = [_dept_to_dict(d) for d in Department.query.filter_by(org_id=caller_org_id).all()]
    else:
        org = Organization.query.get(caller_org_id)
        visible_orgs = [_org_to_dict(org)] if org else []
        dept = Department.query.get(caller_dept_id)
        visible_depts = [_dept_to_dict(dept)] if dept else []
    return render_template("register.html",
        caller_role=caller_role,
        caller_org_id=caller_org_id or "",
        caller_dept_id=caller_dept_id or "",
        visible_orgs=visible_orgs,
        visible_depts=visible_depts
    )

# ─── Page routes: Public token registration ────────────────────────────────────

@app.route("/register/<reg_token>")
def register_token(reg_token):
    """Public, token-gated mobile employee self-registration page (REG-TOKEN-01..02..03)."""
    org = Organization.query.filter_by(reg_token=reg_token).first()
    if not org:
        return render_template("error_token.html", message="Ссылка недействительна"), 404
    org_dict = _org_to_dict(org)
    if is_reg_token_expired(org_dict):
        return render_template("error_token.html",
                               message="Ссылка истекла. Обратитесь к администратору."), 410
    org_id = org.id
    org_depts = [_dept_to_dict(d) for d in Department.query.filter_by(org_id=org_id).all()]
    return render_template(
        "register_token.html",
        reg_token=reg_token,
        org_id=org_id,
        org_name=org.name,
        has_pin=bool(org.reg_pin),
        depts=org_depts,
    )


# ─── API: Public token registration ───────────────────────────────────────────

@app.route("/api/register/<reg_token>/verify_pin", methods=["POST"])
def register_token_verify_pin(reg_token):
    """Verify reg_pin for a registration token link (REG-TOKEN-04..05)."""
    org = Organization.query.filter_by(reg_token=reg_token).first()
    if not org:
        return jsonify({"error": "not_found"}), 404
    if is_reg_token_expired(_org_to_dict(org)):
        return jsonify({"error": "link_expired"}), 410
    stored = org.reg_pin
    if not stored:
        # No PIN configured — open registration
        return jsonify({"verified": True})
    entered = str((request.json or {}).get("pin", ""))
    if len(entered) != 4 or not entered.isdigit():
        return jsonify({"error": "invalid_pin"}), 400
    if bcrypt.checkpw(entered.encode(), stored.encode()):
        return jsonify({"verified": True})
    return jsonify({"error": "wrong_pin", "verified": False}), 401


@app.route("/api/register/<reg_token>/submit", methods=["POST"])
def register_token_submit(reg_token):
    """Create a new employee scoped to the registration token's org (REG-TOKEN-06)."""
    org = Organization.query.filter_by(reg_token=reg_token).first()
    if not org:
        return jsonify({"error": "not_found"}), 404
    org_id = org.id
    if is_reg_token_expired(_org_to_dict(org)):
        return jsonify({"error": "link_expired"}), 410

    data = request.json or {}
    name = data.get("name", "").strip()
    dept_id = data.get("dept_id", "")

    if not name:
        return jsonify({"error": "ФИО обязательно"}), 400

    # Validate dept_id belongs to the token's org
    if dept_id:
        dept = Department.query.get(dept_id)
        if not dept or dept.org_id != org_id:
            return jsonify({"error": "Недопустимый отдел для этой ссылки"}), 400

    emp_id = str(int(time.time() * 1000))
    label = Employee.query.count() + 1

    try:
        emp = Employee(
            id=emp_id,
            name=name,
            role="employee",
            label=label,
            registered_at=datetime.now().isoformat(),
            face_count=0,
            org_id=org_id,
            dept_id=dept_id or None,
        )
        db.session.add(emp)
        db.session.add(EmployeeSchedule(
            emp_id=emp_id,
            start_time="09:00",
            end_time="18:00",
            work_days_json=json.dumps([1, 2, 3, 4, 5]),
        ))
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500

    os.makedirs(os.path.join(FACES_DIR, emp_id), exist_ok=True)
    print(f"REGISTER_TOKEN: emp_id={emp_id!r} name={name!r} org_id={org_id!r} dept_id={dept_id!r}", flush=True)
    # Note: write_audit uses session; self-registration has no session user — actor_username
    # will fall back to None (no admin context). This is expected for public registration.
    write_audit("employee_register", target_type="employee", target_id=emp_id,
                new_value={"name": name, "org_id": org_id, "dept_id": dept_id})
    return jsonify({"id": emp_id, "status": "created"})


@app.route("/api/register/<reg_token>/pending")
def register_token_pending(reg_token):
    """List employees with face_count == 0 for the token's org (new self-registration flow)."""
    org = Organization.query.filter_by(reg_token=reg_token).first()
    if not org:
        return jsonify({"error": "not_found"}), 404
    if is_reg_token_expired(_org_to_dict(org)):
        return jsonify({"error": "link_expired"}), 410
    dept_id = request.args.get("dept_id")
    q = Employee.query.filter_by(org_id=org.id).filter(Employee.face_count == 0)
    if dept_id:
        q = q.filter_by(dept_id=dept_id)
    emps = q.order_by(Employee.name).all()
    return jsonify([{"id": e.id, "name": e.name, "position": e.role or ""} for e in emps])


@app.route("/api/register/<reg_token>/capture_face", methods=["POST"])
def register_token_capture_face(reg_token):
    """Public face photo upload for the self-registration flow — no admin auth required."""
    org = Organization.query.filter_by(reg_token=reg_token).first()
    if not org:
        return jsonify({"error": "not_found"}), 404
    if is_reg_token_expired(_org_to_dict(org)):
        return jsonify({"error": "link_expired"}), 410
    data = request.json or {}
    emp_id = data.get("emp_id")
    image = data.get("image")
    if not emp_id or not image:
        return jsonify({"error": "emp_id and image required"}), 400
    emp = Employee.query.get(emp_id)
    if not emp or emp.org_id != org.id:
        return jsonify({"error": "Сотрудник не найден"}), 404
    img = decode_image(image)
    face_roi, bbox = extract_face(img)
    if face_roi is None:
        return jsonify({"error": "Лицо не обнаружено. Убедитесь, что лицо хорошо видно в кадре."}), 400
    emp_dir = os.path.join(FACES_DIR, emp_id)
    os.makedirs(emp_dir, exist_ok=True)
    count = emp.face_count + 1
    cv2.imwrite(os.path.join(emp_dir, f"face_{count}.jpg"), face_roi)
    emp.face_count = count
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    train_recognizer()
    return jsonify({"status": "saved", "count": count, "bbox": bbox})


ROLE_DISPLAY = {
    "superadmin": "Суперадмин",
    "org_admin": "Администратор организации",
    "dept_admin": "Администратор отдела",
    "hr_viewer": "HR-наблюдатель",
    "viewer": "Наблюдатель",
    "employee": "Сотрудник",
}

@app.route("/admin")
@require_role("superadmin", "org_admin", "dept_admin")
def admin_page():
    if session.get("role") == "superadmin":
        return redirect(url_for("superadmin_page"))
    user = User.query.get(session.get("user_id"))
    username = user.username if user else ""
    creator_role = user.role if user else ""
    creatable_roles = []
    if creator_role in ROLE_HIERARCHY:
        creator_idx = ROLE_HIERARCHY.index(creator_role)
        for role_key in ROLE_HIERARCHY[creator_idx + 1:]:
            creatable_roles.append((role_key, ROLE_DISPLAY.get(role_key, role_key)))
    return render_template("admin.html", username=username, creatable_roles=creatable_roles)

@app.route("/employee")
@require_role("employee")
def employee_page():
    """EMP-01/02/03: Employee self-service cabinet — own T-13 row for current/previous month."""
    user = User.query.get(session.get("user_id"))
    username = user.username if user else ""

    # Resolve emp_id from session user only (T-04-EMP-IDOR: never accept from URL)
    emp_id = user.emp_id if user else None
    if not emp_id:
        return render_template(
            "employee.html",
            username=username,
            emp=None,
            error="Ваш аккаунт не привязан к записи сотрудника. Обратитесь к администратору.",
        )
    emp_obj = Employee.query.get(emp_id)
    if not emp_obj:
        return render_template("403.html"), 403

    # Month clamp (D-08, Pitfall 6): server computes both bounds; client param is untrusted
    now = datetime.now()
    current_month = now.strftime("%Y-%m")
    prev_month = (now.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
    month_str = request.args.get("month", current_month)
    if month_str < prev_month or month_str > current_month:
        month_str = current_month
    year, month_num = map(int, month_str.split("-"))

    # Build days list
    _, num_days = calendar.monthrange(year, month_num)
    days = [date(year, month_num, 1) + timedelta(days=i) for i in range(num_days)]

    # Load attendance records for this employee for the selected month
    start_str = f"{year:04d}-{month_num:02d}-01"
    end_str = f"{year:04d}-{month_num:02d}-{num_days:02d}"
    att_recs = AttendanceRecord.query.filter(
        AttendanceRecord.emp_id == emp_id,
        AttendanceRecord.date >= start_str,
        AttendanceRecord.date <= end_str,
    ).all()

    # Build both the attendance dict (for compute_symbol) and times_by_date (for tooltips)
    attendance = {}
    times_by_date = {}
    for r in att_recs:
        attendance.setdefault(r.date, {})[r.emp_id] = {
            "check_in": r.check_in_time,
            "check_out": r.check_out_time,
        }
        times_by_date[r.date] = {
            "check_in": r.check_in_time,
            "check_out": r.check_out_time,
        }

    # Load overrides for this employee
    ov_recs = TimesheetOverride.query.filter_by(emp_id=emp_id).all()
    overrides = {emp_id: {r.date: r.symbol for r in ov_recs}}

    holidays_set = get_holidays_set(year)

    # Build single-row grid (same inline per-cell loop as /timesheet)
    emp_dict = _emp_to_dict(emp_obj)
    schedule = emp_dict.get("schedule", {"start": "09:00", "end": "18:00", "work_days": [1, 2, 3, 4, 5]})
    cells = []
    for d in days:
        sym = compute_symbol(d, emp_id, attendance, overrides, schedule, holidays_set)
        auto = compute_symbol(d, emp_id, attendance, {}, schedule, holidays_set)
        cells.append({"sym": sym, "auto": auto, "date": d.isoformat()})

    symbols = [c["sym"] for c in cells]
    totals = compute_employee_totals(symbols, schedule)
    grid_row = (emp_id, emp_obj.name, cells, totals)

    # Stats (EMP-03): early_count from У/ОУ symbols (not returned by compute_employee_totals)
    early_count = sum(1 for s in symbols if s in ("У", "ОУ"))
    stats = {
        "late": totals["late"],
        "absences": totals["absences"],
        "early": early_count,
    }

    return render_template(
        "employee.html",
        username=username,
        emp_name=emp_obj.name,
        grid_row=grid_row,
        stats=stats,
        times_by_date=times_by_date,
        days=days,
        month_str=month_str,
        current_month=current_month,
        prev_month=prev_month,
        holidays_set=holidays_set,
        error=None,
    )

# ─── Page routes: Role Dashboards ─────────────────────────────────────────────

@app.route("/superadmin/<tab>")
@app.route("/superadmin")
@require_role("superadmin")
def superadmin_page(tab="orgs"):
    VALID_TABS = {"orgs", "users"}
    initial_tab = tab if tab in VALID_TABS else "orgs"
    user = User.query.get(session.get("user_id"))
    username = user.username if user else ""
    role = user.role if user else ""
    return render_template("superadmin.html", username=username, role=role, initial_tab=initial_tab)


@app.route("/org_admin/<tab>")
@app.route("/org_admin")
@require_role("org_admin")
def org_admin_page(tab="depts"):
    VALID_TABS = {"depts", "employees", "summary", "reports", "users", "settings", "timesheet"}
    initial_tab = tab if tab in VALID_TABS else "depts"
    user = User.query.get(session.get("user_id"))
    username = user.username if user else ""
    role = user.role if user else ""
    caller_org_id = session.get("org_id")
    org = Organization.query.get(caller_org_id)
    org_name = org.name if org else ""
    org_token = org.org_token or "" if org else ""
    reg_token = org.reg_token or "" if org else ""
    reg_token_expires = org.reg_token_expires if org else None
    kiosk_display_name = org.kiosk_display_name or "" if org else ""

    # DASH-04: optional ?summary_month=YYYY-MM — compute per-dept summary when present
    # Scope is always session["org_id"] — never from client (T-03-summary-scope)
    summary_month = request.args.get("summary_month", "")
    summary_rows = []
    if summary_month:
        try:
            sum_year, sum_month_num = map(int, summary_month.split("-"))
            if not (1 <= sum_month_num <= 12 and 2000 <= sum_year <= 2100):
                raise ValueError("out of range")
            import calendar as _cal
            from datetime import date as _date, timedelta as _td
            _start = f"{sum_year:04d}-{sum_month_num:02d}-01"
            _, _nd = _cal.monthrange(sum_year, sum_month_num)
            _end = f"{sum_year:04d}-{sum_month_num:02d}-{_nd:02d}"
            # Load data for summary computation
            employees = {e.id: _emp_to_dict(e) for e in Employee.query.all()}
            # Build attendance dict adapter from ORM
            _att_recs = AttendanceRecord.query.filter(
                AttendanceRecord.date >= _start,
                AttendanceRecord.date <= _end,
            ).all()
            attendance = {}
            for r in _att_recs:
                attendance.setdefault(r.date, {})[r.emp_id] = {
                    "check_in": r.check_in_time,
                    "check_out": r.check_out_time,
                }
            # Build overrides dict adapter from ORM
            _ov_recs = TimesheetOverride.query.all()
            overrides = {}
            for r in _ov_recs:
                overrides.setdefault(r.emp_id, {})[r.date] = r.symbol
            rows = compute_dept_summary(
                sum_year, sum_month_num, caller_org_id, employees, attendance, overrides
            )
            # Enrich dept_name from Department ORM
            for row in rows:
                dept = Department.query.get(row["dept_id"])
                if dept:
                    row["dept_name"] = dept.name
            summary_rows = rows
        except (ValueError, AttributeError, TypeError):
            summary_month = ""
            summary_rows = []

    return render_template(
        "org_admin.html",
        username=username,
        role=role,
        org_name=org_name,
        org_id=caller_org_id or "",
        org_token=org_token,
        reg_token=reg_token,
        reg_token_expires=reg_token_expires or "",
        kiosk_display_name=kiosk_display_name,
        summary_month=summary_month,
        summary_rows=summary_rows,
        initial_tab=initial_tab,
    )


@app.route("/dept_admin/<tab>")
@app.route("/dept_admin")
@require_role("dept_admin", "viewer")
def dept_admin_page(tab="attendance"):
    VALID_TABS = {"attendance", "employees", "timesheet"}
    initial_tab = tab if tab in VALID_TABS else "attendance"
    user = User.query.get(session.get("user_id"))
    username = user.username if user else ""
    role = user.role if user else ""
    dept = Department.query.get(session.get("dept_id"))
    dept_name = dept.name if dept else ""
    return render_template("dept_admin.html", username=username, role=role, dept_name=dept_name, initial_tab=initial_tab)


@app.route("/dashboard")
@require_role()
def dashboard_page():
    user = User.query.get(session.get("user_id"))
    username = user.username if user else ""
    return render_template("dashboard.html", username=username)

@app.route("/profile", methods=["GET", "POST"])
@require_role()
def profile_page():
    error = None
    success = None
    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")
        user_id = session.get("user_id")
        user = User.query.get(user_id)
        if not user or not bcrypt.checkpw(current_password.encode(), user.password_hash.encode()):
            error = "Текущий пароль введён неверно"
        elif new_password != confirm_password:
            error = "Новые пароли не совпадают"
        elif len(new_password) < 8:
            error = "Пароль должен содержать не менее 8 символов"
        else:
            try:
                user.password_hash = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
                db.session.commit()
                success = "Пароль успешно изменён"
            except Exception:
                db.session.rollback()
                error = "Internal server error"
    return render_template("profile.html", error=error, success=success)


@app.route("/account")
@require_role()
def account_page():
    user = User.query.get(session.get("user_id"))
    username = user.username if user else ""
    display_name = (user.display_name or "") if user else ""
    return render_template("account.html", username=username, display_name=display_name)


@app.route("/api/me", methods=["PATCH"])
@require_role()
def api_me():
    user = User.query.get(session.get("user_id"))
    if not user:
        return jsonify({"error": "Не авторизовано"}), 401
    data = request.get_json(silent=True) or {}
    if "display_name" in data:
        name = data["display_name"].strip()
        if len(name) > 128:
            return jsonify({"error": "Имя слишком длинное"}), 400
        user.display_name = name or None
    if "new_password" in data:
        current_password = data.get("current_password", "")
        new_password = data["new_password"]
        if not bcrypt.checkpw(current_password.encode(), user.password_hash.encode()):
            return jsonify({"error": "Текущий пароль введён неверно"}), 400
        if len(new_password) < 8:
            return jsonify({"error": "Пароль должен содержать не менее 8 символов"}), 400
        user.password_hash = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    return jsonify({"status": "updated", "display_name": user.display_name or ""})


# ─── Page routes: T-13 Timesheet ──────────────────────────────────────────────

@app.route("/timesheet")
@require_role("dept_admin", "org_admin", "hr_viewer", "superadmin")
def timesheet():
    """T13-01: Render the T-13 attendance timesheet grid for a dept/month."""
    user = User.query.get(session.get("user_id"))
    username = user.username if user else ""
    role = session.get("role")
    session_dept_id = session.get("dept_id")
    session_org_id = session.get("org_id")

    # (a) Resolve month param
    month_str = request.args.get("month", datetime.now().strftime("%Y-%m"))
    try:
        year, month_num = map(int, month_str.split("-"))
        if not (1 <= month_num <= 12 and 2000 <= year <= 2099):
            raise ValueError("out of range")
    except (ValueError, AttributeError):
        year, month_num = datetime.now().year, datetime.now().month
        month_str = f"{year:04d}-{month_num:02d}"

    # (b) Resolve dept scope (D-08: dept_admin param is ignored — always forced to session dept)
    dept_id_param = request.args.get("dept_id", "")
    dept_options = []

    if role == "dept_admin":
        dept_id = session_dept_id  # always fixed from session; param ignored
    elif role in ("org_admin", "hr_viewer"):
        if dept_id_param:
            dept_obj = Department.query.get(dept_id_param)
            if not dept_obj or dept_obj.org_id != session_org_id:
                return render_template("403.html"), 403
            dept_id = dept_id_param
        else:
            # Default to first dept in org
            first_dept = Department.query.filter_by(org_id=session_org_id).first()
            dept_id = first_dept.id if first_dept else None
        # Build dept_options for selector
        dept_options = [
            {"id": d.id, "name": d.name}
            for d in Department.query.filter_by(org_id=session_org_id).all()
        ]
    else:  # superadmin
        dept_id = dept_id_param or None
        # Build dept_options grouped by org for superadmin
        dept_options = []
        for org_obj in Organization.query.all():
            org_depts = [
                {"id": d.id, "name": d.name, "org_name": org_obj.name}
                for d in Department.query.filter_by(org_id=org_obj.id).all()
            ]
            if org_depts:
                dept_options.append({"org_id": org_obj.id, "org_name": org_obj.name, "depts": org_depts})

    # Resolve dept name for display
    if dept_id:
        dept_obj = Department.query.get(dept_id)
        dept_name = dept_obj.name if dept_obj else ""
    else:
        dept_name = ""

    # (c) Build days list
    _, num_days = calendar.monthrange(year, month_num)
    days = [date(year, month_num, 1) + timedelta(days=i) for i in range(num_days)]

    # (d) Load data via ORM
    employees = {e.id: _emp_to_dict(e) for e in Employee.query.all()}
    # Build attendance dict adapter: {date: {emp_id: {check_in, check_out}}}
    start_str = f"{year:04d}-{month_num:02d}-01"
    _, _num_days_ts = calendar.monthrange(year, month_num)
    end_str = f"{year:04d}-{month_num:02d}-{_num_days_ts:02d}"
    _att_recs = AttendanceRecord.query.filter(
        AttendanceRecord.date >= start_str,
        AttendanceRecord.date <= end_str,
    ).all()
    attendance = {}
    for r in _att_recs:
        attendance.setdefault(r.date, {})[r.emp_id] = {
            "check_in": r.check_in_time,
            "check_out": r.check_out_time,
        }
    # Build overrides dict adapter: {emp_id: {date: symbol}}
    _ov_recs = TimesheetOverride.query.all()
    overrides = {}
    for r in _ov_recs:
        overrides.setdefault(r.emp_id, {})[r.date] = r.symbol
    holidays_set = get_holidays_set(year)
    missing_holiday_year = is_holiday_year_missing(year)

    # Filter employees to dept scope (Pitfall 3: None dept_id is out-of-scope)
    if dept_id:
        scoped_employees = {eid: e for eid, e in employees.items() if e.get("dept_id") == dept_id}
    else:
        scoped_employees = {}

    # (e) Build grid rows and totals
    # Each cell is a dict: {sym: displayed_symbol, auto: auto_symbol, date: iso_date_str}
    # auto is the symbol computed without overrides — used by "Восстановить автоматически"
    # to repaint the cell client-side without a page reload.
    grid_rows = []
    for emp_id, emp in scoped_employees.items():
        schedule = emp.get("schedule", {"start": "09:00", "end": "18:00", "work_days": [1, 2, 3, 4, 5]})
        cells = []
        for d in days:
            sym = compute_symbol(d, emp_id, attendance, overrides, schedule, holidays_set)
            auto = compute_symbol(d, emp_id, attendance, {}, schedule, holidays_set)
            cells.append({"sym": sym, "auto": auto, "date": d.isoformat()})
        # symbols list for totals computation: use displayed symbols (overrides included)
        symbols = [c["sym"] for c in cells]
        totals = compute_employee_totals(symbols, schedule)
        grid_rows.append((emp_id, emp.get("name", emp_id), cells, totals))

    # (f) Render template
    return render_template(
        "timesheet.html",
        username=username,
        role=role,
        dept_name=dept_name,
        dept_id=dept_id,
        dept_options=dept_options,
        month_str=month_str,
        year=year,
        month_num=month_num,
        days=days,
        grid_rows=grid_rows,
        holidays_set=holidays_set,
        missing_holiday_year=missing_holiday_year,
        can_edit=(role in ("dept_admin", "org_admin", "superadmin")),
        # hr_viewer is read-only; can_edit stays False for them
    )


# ─── Export: T-13 XLSX / CSV ──────────────────────────────────────────────────

def _resolve_export_scope():
    """Shared scope/data resolution for export routes (mirrors /timesheet logic).

    Returns (dept_id, dept_name, org_name, year, month_num, month_str, days,
             scoped_employees, attendance, overrides, holidays_set) or raises a
    Flask response tuple (status_code, response) that the caller should return.
    """
    role = session.get("role")
    session_dept_id = session.get("dept_id")
    session_org_id = session.get("org_id")

    # (a) Resolve month param (T-04-VAL)
    month_str = request.args.get("month", datetime.now().strftime("%Y-%m"))
    try:
        year, month_num = map(int, month_str.split("-"))
        if not (1 <= month_num <= 12 and 2000 <= year <= 2099):
            raise ValueError("out of range")
    except (ValueError, AttributeError):
        year, month_num = datetime.now().year, datetime.now().month
        month_str = f"{year:04d}-{month_num:02d}"

    # (b) Resolve dept scope (T-04-IDOR — copy verbatim from /timesheet)
    dept_id_param = request.args.get("dept_id", "")

    if role == "dept_admin":
        dept_id = session_dept_id  # always fixed from session; param ignored
    elif role in ("org_admin", "hr_viewer"):
        if dept_id_param:
            dept_obj = Department.query.get(dept_id_param)
            if not dept_obj or dept_obj.org_id != session_org_id:
                return None, (render_template("403.html"), 403)
            dept_id = dept_id_param
        else:
            first_dept = Department.query.filter_by(org_id=session_org_id).first()
            dept_id = first_dept.id if first_dept else None
    else:  # superadmin
        dept_id = dept_id_param or None

    # Guard: no dept selected
    if not dept_id:
        return None, (render_template("403.html"), 403)

    # Resolve dept/org names for header rows
    dept_obj = Department.query.get(dept_id)
    dept_name = dept_obj.name if dept_obj else dept_id
    org_obj = Organization.query.get(dept_obj.org_id) if dept_obj else None
    org_name = org_obj.name if org_obj else ""

    # (c) Build days list
    _, num_days = calendar.monthrange(year, month_num)
    days = [date(year, month_num, 1) + timedelta(days=i) for i in range(num_days)]

    # (d) Load data via ORM (mirrors /timesheet lines 933-953)
    employees = {e.id: _emp_to_dict(e) for e in Employee.query.all()}
    start_str = f"{year:04d}-{month_num:02d}-01"
    end_str = f"{year:04d}-{month_num:02d}-{num_days:02d}"
    _att_recs = AttendanceRecord.query.filter(
        AttendanceRecord.date >= start_str,
        AttendanceRecord.date <= end_str,
    ).all()
    attendance = {}
    for r in _att_recs:
        attendance.setdefault(r.date, {})[r.emp_id] = {
            "check_in": r.check_in_time,
            "check_out": r.check_out_time,
        }
    _ov_recs = TimesheetOverride.query.all()
    overrides = {}
    for r in _ov_recs:
        overrides.setdefault(r.emp_id, {})[r.date] = r.symbol
    holidays_set = get_holidays_set(year)

    # Filter to dept scope (Pitfall 3)
    scoped_employees = {eid: e for eid, e in employees.items() if e.get("dept_id") == dept_id}

    ctx = (dept_id, dept_name, org_name, year, month_num, month_str, days,
           scoped_employees, attendance, overrides, holidays_set)
    return ctx, None


def _build_export_grid(days, scoped_employees, attendance, overrides, holidays_set):
    """Build grid rows as (emp_id, name, symbols, totals) for export routes."""
    grid_rows = []
    for emp_id, emp in scoped_employees.items():
        schedule = emp.get("schedule", {"start": "09:00", "end": "18:00", "work_days": [1, 2, 3, 4, 5]})
        symbols = [compute_symbol(d, emp_id, attendance, overrides, schedule, holidays_set) for d in days]
        totals = compute_employee_totals(symbols, schedule)
        grid_rows.append((emp_id, emp.get("name", emp_id), symbols, totals))
    return grid_rows


@app.route("/timesheet/export/xlsx")
@require_role("dept_admin", "org_admin", "hr_viewer", "superadmin")
def export_timesheet_xlsx():
    """EXP-01: Download the T-13 grid as an openpyxl .xlsx file."""
    ctx, err = _resolve_export_scope()
    if err is not None:
        return err
    dept_id, dept_name, org_name, year, month_num, month_str, days, scoped_employees, attendance, overrides, holidays_set = ctx

    grid_rows = _build_export_grid(days, scoped_employees, attendance, overrides, holidays_set)

    wb = Workbook()
    ws = wb.active

    num_cols = 1 + len(days) + 5  # name + day columns + 5 total columns

    # Row 1: merged title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value="ТАБЕЛЬ УЧЁТА РАБОЧЕГО ВРЕМЕНИ (Форма Т-13)")
    title_cell.font = Font(bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # Row 2: merged subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    sub_cell = ws.cell(row=2, column=1, value=f"{org_name} — {dept_name} — {MONTHS_RU[month_num]} {year}")
    sub_cell.alignment = Alignment(horizontal="center")

    # Row 3: headers
    headers = ["Сотрудник"] + [d.day for d in days] + ["Я", "Ч", "П/НН", "О", "Б/К"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Rows 4+: employee data
    for row_idx, (emp_id, name, symbols, totals) in enumerate(grid_rows, start=4):
        ws.cell(row=row_idx, column=1, value=name)
        for col_offset, sym in enumerate(symbols, start=2):
            ws.cell(row=row_idx, column=col_offset, value=sym if sym else "")
        # 5 totals columns
        base_col = 2 + len(days)
        ws.cell(row=row_idx, column=base_col, value=totals["days_worked"])
        ws.cell(row=row_idx, column=base_col + 1, value=totals["hours_worked"])
        ws.cell(row=row_idx, column=base_col + 2, value=totals["absences"])
        ws.cell(row=row_idx, column=base_col + 3, value=totals["late"])
        ws.cell(row=row_idx, column=base_col + 4, value=totals["vac_sick"])

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 24
    for col_idx in range(2, 2 + len(days)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 4

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)  # Pitfall 2: must seek before send_file

    safe_dept = re.sub(r"[^A-Za-zА-Яа-яЁё0-9]", "_", dept_name)
    return send_file(
        buf,
        download_name=f"T13_{safe_dept}_{month_str}.xlsx",
        as_attachment=True,
    )


@app.route("/timesheet/export/csv")
@require_role("dept_admin", "org_admin", "hr_viewer", "superadmin")
def export_timesheet_csv():
    """EXP-02: Download the T-13 grid as UTF-8 BOM semicolon-delimited CSV."""
    ctx, err = _resolve_export_scope()
    if err is not None:
        return err
    dept_id, dept_name, org_name, year, month_num, month_str, days, scoped_employees, attendance, overrides, holidays_set = ctx

    grid_rows = _build_export_grid(days, scoped_employees, attendance, overrides, holidays_set)

    str_buf = io.StringIO()
    writer = csv.writer(str_buf, delimiter=";")

    # Title row
    writer.writerow([f"ТАБЕЛЬ Т-13 — {dept_name} — {month_str}"])
    # Header row
    headers = ["Сотрудник"] + [str(d.day) for d in days] + ["Я", "Ч", "П/НН", "О", "Б/К"]
    writer.writerow(headers)
    # Employee rows
    for emp_id, name, symbols, totals in grid_rows:
        row = [name] + [sym if sym else "" for sym in symbols] + [
            totals["days_worked"],
            totals["hours_worked"],
            totals["absences"],
            totals["late"],
            totals["vac_sick"],
        ]
        writer.writerow(row)

    content = str_buf.getvalue().encode("utf-8-sig")  # BOM via codec (Pitfall: do NOT hand-write BOM)
    byte_buf = BytesIO(content)

    safe_dept = re.sub(r"[^A-Za-zА-Яа-яЁё0-9]", "_", dept_name)
    return send_file(
        byte_buf,
        download_name=f"T13_{safe_dept}_{month_str}.csv",
        as_attachment=True,
        mimetype="text/csv",
    )


# ─── API: Manual attendance time edit ────────────────────────────────────────

@app.route("/api/attendance/manual", methods=["POST"])
@require_role("superadmin", "org_admin")
def manual_time_edit():
    """Manually set check_in or check_out for an employee on a date.

    Allowed roles: superadmin (global), org_admin (own org only).
    dept_admin and hr_viewer → 403 (not listed in require_role).
    Writes an AuditLog row after successful commit.
    """
    data = request.get_json(silent=True) or {}
    emp_id = data.get("emp_id", "")
    date_str = data.get("date", "")
    check_in_raw = data.get("check_in")   # HH:MM or HH:MM:SS or empty/null
    check_out_raw = data.get("check_out")  # same

    # Validate emp exists
    emp = Employee.query.get(emp_id)
    if not emp:
        return jsonify({"error": "employee_not_found"}), 404

    # Scope gate: org_admin may only edit own org
    role = session.get("role")
    if role == "org_admin" and emp.org_id != session.get("org_id"):
        return jsonify({"error": "forbidden"}), 403

    # Validate date
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
    except (ValueError, TypeError):
        return jsonify({"error": "invalid_date", "message": "Ожидается YYYY-MM-DD"}), 422

    # Validate and normalize each time value
    def _normalize_time(raw):
        """Return HH:MM:SS string, None (clear), or raise ValueError on bad input."""
        if raw is None or raw == "":
            return None  # clear the field
        s = str(raw).strip()
        if not re.match(r'^\d{2}:\d{2}(:\d{2})?$', s):
            raise ValueError(f"invalid time format: {s!r}")
        parts = s.split(":")
        h, m = int(parts[0]), int(parts[1])
        sec = int(parts[2]) if len(parts) == 3 else 0
        if not (0 <= h <= 23 and 0 <= m <= 59 and 0 <= sec <= 59):
            raise ValueError(f"time out of range: {s!r}")
        return f"{h:02d}:{m:02d}:{sec:02d}"

    try:
        new_ci = _normalize_time(check_in_raw)
        new_co = _normalize_time(check_out_raw)
    except ValueError as e:
        return jsonify({"error": "invalid_time", "message": str(e)}), 422

    # If no meaningful changes requested, bail early
    if check_in_raw is None and check_out_raw is None:
        return jsonify({"error": "no_fields"}), 422

    # Upsert AttendanceRecord
    rec = AttendanceRecord.query.filter_by(emp_id=emp_id, date=date_str).first()
    old_ci = rec.check_in_time if rec else None
    old_co = rec.check_out_time if rec else None

    try:
        if rec is None:
            rec = AttendanceRecord(
                emp_id=emp_id,
                date=date_str,
                check_in_time=new_ci,
                check_out_time=new_co,
                event_type="manual",
            )
            db.session.add(rec)
        else:
            if check_in_raw is not None:
                rec.check_in_time = new_ci
            if check_out_raw is not None:
                rec.check_out_time = new_co
            rec.event_type = "manual"
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500

    write_audit("manual_time_edit", target_type="attendance",
                target_id=f"{emp_id}:{date_str}",
                old_value={"check_in": old_ci, "check_out": old_co},
                new_value={"check_in": rec.check_in_time, "check_out": rec.check_out_time})

    return jsonify({
        "status": "updated",
        "check_in": rec.check_in_time,
        "check_out": rec.check_out_time,
    })


# ─── API: T-13 Timesheet Override ────────────────────────────────────────────

@app.route("/api/timesheet/override", methods=["POST", "DELETE"])
@require_role("dept_admin", "org_admin", "superadmin")
def timesheet_override():
    """D-05: Inline manual override for timesheet cells.

    POST: set a manual symbol (Б/К/П) for emp_id on date.
    DELETE: remove override and restore auto-derived symbol.
    Scope-checked server-side from employees.json (never trust client dept/org).
    """
    role = session.get("role")
    session_dept_id = session.get("dept_id")
    session_org_id = session.get("org_id")

    data = request.get_json(silent=True) or {}
    emp_id = data.get("emp_id", "")
    date_str = data.get("date", "")

    # Validate emp exists
    emp = Employee.query.get(emp_id)
    if not emp:
        return jsonify({"error": "employee_not_found"}), 404

    # Scope check: read dept/org from employee record (T-03-privesc mitigation)
    if role == "dept_admin" and emp.dept_id != session_dept_id:
        return jsonify({"error": "forbidden"}), 403
    if role == "org_admin" and emp.org_id != session_org_id:
        return jsonify({"error": "forbidden"}), 403
    # superadmin: unrestricted

    if request.method == "DELETE":
        try:
            existing = db.session.get(TimesheetOverride, (emp_id, date_str))
            if existing:
                db.session.delete(existing)
                db.session.commit()
        except Exception:
            db.session.rollback()
            return jsonify({"error": "Internal server error"}), 500
        return jsonify({"deleted": True})

    # POST branch
    symbol = data.get("symbol", "")

    # Validate symbol is in the manual whitelist (T-03-inject mitigation)
    if symbol not in MANUAL_SYMBOLS:
        return jsonify({"error": "invalid_symbol"}), 422

    # Validate date_str is a real YYYY-MM-DD calendar date (rejects "2025-13-01", "aaaa-bb-cc", etc.)
    try:
        datetime.strptime(date_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return jsonify({"error": "invalid_date"}), 422

    try:
        updated_by = session.get("user_id")
        updated_at = datetime.now().isoformat()
        existing = db.session.get(TimesheetOverride, (emp_id, date_str))
        if existing:
            existing.symbol = symbol
            existing.updated_by = updated_by
            existing.updated_at = updated_at
        else:
            db.session.add(TimesheetOverride(
                emp_id=emp_id,
                date=date_str,
                symbol=symbol,
                updated_by=updated_by,
                updated_at=updated_at,
            ))
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    return jsonify({"symbol": symbol, "auto": False})


# ─── API: Users ───────────────────────────────────────────────────────────────

@app.route("/api/users", methods=["GET"])
@require_role("superadmin", "org_admin", "dept_admin")
def list_users():
    caller_role = session.get("role")
    caller_org_id = session.get("org_id")
    if caller_role == "org_admin":
        all_users = User.query.filter_by(org_id=caller_org_id).all()
    else:
        all_users = User.query.all()
    result = [
        {
            "id": u.id,
            "username": u.username,
            "role": u.role,
            "active": u.active,
            "org_id": u.org_id,
            "dept_id": u.dept_id,
        }
        for u in all_users
    ]
    return jsonify(result)

@app.route("/api/users", methods=["POST"])
@require_role("superadmin", "org_admin", "dept_admin")
def create_user():
    data = request.get_json(silent=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    target_role = data.get("role", "")
    creator_role = session.get("role")
    if not username:
        return jsonify({"error": "Логин не может быть пустым"}), 400
    if len(password) < 8:
        return jsonify({"error": "Пароль должен содержать не менее 8 символов"}), 400
    if target_role not in ROLE_HIERARCHY:
        return jsonify({"error": "Недопустимая роль"}), 400
    if target_role == "viewer":
        return jsonify({"error": "Роль 'viewer' не поддерживается в текущей версии"}), 400
    if (creator_role not in ROLE_HIERARCHY or
            ROLE_HIERARCHY.index(creator_role) >= ROLE_HIERARCHY.index(target_role)):
        return jsonify({"error": "forbidden"}), 403
    # superadmin may only create org_admin; org_admin manages all roles below them
    if creator_role == "superadmin" and target_role != "org_admin":
        return jsonify({"error": "Суперадминистратор может создавать только администраторов организаций"}), 403
    if User.query.filter_by(username=username).first():
        return jsonify({"error": "Пользователь с таким логином уже существует"}), 400
    # Determine org scope: org_admin and dept_admin are always forced to their own org
    caller_org_id = session.get("org_id")
    caller_dept_id = session.get("dept_id")
    if creator_role == "superadmin":
        new_org_id = data.get("org_id")
    else:
        new_org_id = caller_org_id  # org_admin/dept_admin may never cross org boundary
    new_dept_id = data.get("dept_id") or (caller_dept_id if creator_role == "dept_admin" else None)
    # emp_id only applies to employee-role accounts; force None for all other roles (T-04-EMP-LINK)
    new_emp_id = (data.get("emp_id") or None) if target_role == "employee" else None
    user_id = str(uuid.uuid4())
    try:
        db.session.add(User(
            id=user_id,
            username=username,
            password_hash=bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode(),
            role=target_role,
            active=True,
            org_id=new_org_id,
            dept_id=new_dept_id,
            emp_id=new_emp_id,
        ))
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    write_audit("user_create", target_type="user", target_id=user_id,
                new_value={"username": username, "role": target_role, "org_id": new_org_id, "dept_id": new_dept_id})
    print(f"USER_CREATED: username={username!r} role={target_role!r} org_id={new_org_id!r} dept_id={new_dept_id!r} emp_id={new_emp_id!r}", flush=True)
    return jsonify({"id": user_id, "status": "created"})

@app.route("/api/users/<user_id>", methods=["PATCH"])
@require_role("superadmin", "org_admin", "dept_admin")
def update_user(user_id):
    target = User.query.get(user_id)
    if not target:
        return jsonify({"error": "Пользователь не найден"}), 404
    caller_role = session.get("role")
    target_role = target.role
    if (caller_role not in ROLE_HIERARCHY or
            target_role not in ROLE_HIERARCHY or
            ROLE_HIERARCHY.index(caller_role) >= ROLE_HIERARCHY.index(target_role)):
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json(silent=True) or {}
    old_active = target.active
    old_dept_id = target.dept_id
    changes = {}
    if "active" in data:
        target.active = bool(data["active"])
        changes["active"] = data["active"]
    if "password" in data:
        pw = data["password"]
        if len(pw) < 8:
            return jsonify({"error": "Пароль должен содержать не менее 8 символов"}), 400
        target.password_hash = bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()
        changes["password"] = "***"
    if "dept_id" in data:
        target.dept_id = data["dept_id"] or None
        changes["dept_id"] = data["dept_id"] or None
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    write_audit("user_update", target_type="user", target_id=user_id,
                old_value={"active": old_active, "dept_id": old_dept_id},
                new_value=changes)
    return jsonify({"status": "updated", "active": target.active})


@app.route("/api/users/<user_id>", methods=["DELETE"])
@require_role("superadmin", "org_admin")
def delete_user(user_id):
    target = User.query.get(user_id)
    if not target:
        return jsonify({"error": "Пользователь не найден"}), 404
    caller_role = session.get("role")
    caller_id = session.get("user_id")
    if str(caller_id) == str(user_id):
        return jsonify({"error": "Нельзя удалить собственный аккаунт"}), 403
    if (caller_role not in ROLE_HIERARCHY or
            target.role not in ROLE_HIERARCHY or
            ROLE_HIERARCHY.index(caller_role) >= ROLE_HIERARCHY.index(target.role)):
        return jsonify({"error": "forbidden"}), 403
    if caller_role == "org_admin" and target.org_id != session.get("org_id"):
        return jsonify({"error": "forbidden"}), 403
    old_username = target.username
    old_role = target.role
    try:
        db.session.delete(target)
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    write_audit("user_delete", target_type="user", target_id=user_id,
                old_value={"username": old_username, "role": old_role})
    return jsonify({"status": "deleted"})


# ─── Page routes: Org admin inline partials ───────────────────────────────────

@app.route("/org_admin/partial/reports")
@require_role("org_admin", "hr_viewer", "superadmin")
def org_admin_partial_reports():
    """ORGUX-05: Return reports HTML fragment for inline injection into org_admin."""
    user = User.query.get(session.get("user_id"))
    username = user.username if user else ""
    return render_template("reports_partial.html", username=username)


@app.route("/org_admin/partial/timesheet")
@require_role("org_admin", "hr_viewer", "superadmin")
def org_admin_partial_timesheet():
    """ORGUX-05: Return timesheet HTML fragment for inline injection into org_admin."""
    user = User.query.get(session.get("user_id"))
    username = user.username if user else ""
    role = session.get("role")
    session_org_id = session.get("org_id")

    # Resolve month
    month_str = request.args.get("month", datetime.now().strftime("%Y-%m"))
    try:
        year, month_num = map(int, month_str.split("-"))
        if not (1 <= month_num <= 12 and 2000 <= year <= 2099):
            raise ValueError
    except (ValueError, AttributeError):
        year, month_num = datetime.now().year, datetime.now().month
        month_str = f"{year:04d}-{month_num:02d}"

    # Resolve dept
    dept_id_param = request.args.get("dept_id", "")
    dept_options = []

    if role in ("org_admin", "hr_viewer"):
        if dept_id_param:
            dept_obj = Department.query.get(dept_id_param)
            if not dept_obj or dept_obj.org_id != session_org_id:
                return jsonify({"error": "forbidden"}), 403
            dept_id = dept_id_param
        else:
            first_dept = Department.query.filter_by(org_id=session_org_id).first()
            dept_id = first_dept.id if first_dept else None
        dept_options = [
            {"id": d.id, "name": d.name}
            for d in Department.query.filter_by(org_id=session_org_id).all()
        ]
    else:  # superadmin
        dept_id = dept_id_param or None
        for org_obj in Organization.query.all():
            org_depts = [
                {"id": d.id, "name": d.name, "org_name": org_obj.name}
                for d in Department.query.filter_by(org_id=org_obj.id).all()
            ]
            if org_depts:
                dept_options.append({"org_id": org_obj.id, "org_name": org_obj.name, "depts": org_depts})

    # Resolve dept name
    if dept_id:
        dept_obj = Department.query.get(dept_id)
        dept_name = dept_obj.name if dept_obj else ""
    else:
        dept_name = ""

    # Build days list
    _, num_days = calendar.monthrange(year, month_num)
    days = [date(year, month_num, 1) + timedelta(days=i) for i in range(num_days)]

    # Load data via ORM
    employees = {e.id: _emp_to_dict(e) for e in Employee.query.all()}
    start_str = f"{year:04d}-{month_num:02d}-01"
    _, _nd = calendar.monthrange(year, month_num)
    end_str = f"{year:04d}-{month_num:02d}-{_nd:02d}"
    _att_recs = AttendanceRecord.query.filter(
        AttendanceRecord.date >= start_str,
        AttendanceRecord.date <= end_str,
    ).all()
    attendance = {}
    for r in _att_recs:
        attendance.setdefault(r.date, {})[r.emp_id] = {
            "check_in": r.check_in_time,
            "check_out": r.check_out_time,
        }
    _ov_recs = TimesheetOverride.query.all()
    overrides = {}
    for r in _ov_recs:
        overrides.setdefault(r.emp_id, {})[r.date] = r.symbol
    holidays_set = get_holidays_set(year)
    missing_holiday_year = is_holiday_year_missing(year)

    # Filter to dept scope
    if dept_id:
        scoped_employees = {eid: e for eid, e in employees.items() if e.get("dept_id") == dept_id}
    else:
        scoped_employees = {}

    # Build grid rows
    grid_rows = []
    for emp_id_key, emp in scoped_employees.items():
        schedule = emp.get("schedule", {"start": "09:00", "end": "18:00", "work_days": [1, 2, 3, 4, 5]})
        cells = []
        for d in days:
            sym = compute_symbol(d, emp_id_key, attendance, overrides, schedule, holidays_set)
            auto = compute_symbol(d, emp_id_key, attendance, {}, schedule, holidays_set)
            cells.append({"sym": sym, "auto": auto, "date": d.isoformat()})
        symbols = [c["sym"] for c in cells]
        totals = compute_employee_totals(symbols, schedule)
        grid_rows.append((emp_id_key, emp.get("name", emp_id_key), cells, totals))

    month_name = MONTHS_RU.get(month_num, "")
    return render_template(
        "timesheet_partial.html",
        username=username,
        role=role,
        dept_id=dept_id or "",
        dept_name=dept_name,
        dept_options=dept_options,
        month_str=month_str,
        year=year,
        month_num=month_num,
        month_name=month_name,
        days=days,
        grid_rows=grid_rows,
        holidays_set=holidays_set,
        missing_holiday_year=missing_holiday_year,
        can_edit=(role in ("dept_admin", "org_admin", "superadmin")),
        partial_url="/org_admin/partial/timesheet",
        content_el_id="inlineTimesheetContent",
    )


@app.route("/dept_admin/partial/timesheet")
@require_role("dept_admin")
def dept_admin_partial_timesheet():
    """Return T-13 timesheet fragment for dept_admin, scoped to their department."""
    user = User.query.get(session.get("user_id"))
    username = user.username if user else ""
    dept_id = session.get("dept_id")

    month_str = request.args.get("month", datetime.now().strftime("%Y-%m"))
    try:
        year, month_num = map(int, month_str.split("-"))
        if not (1 <= month_num <= 12 and 2000 <= year <= 2099):
            raise ValueError
    except (ValueError, AttributeError):
        year, month_num = datetime.now().year, datetime.now().month
        month_str = f"{year:04d}-{month_num:02d}"

    dept_obj = Department.query.get(dept_id) if dept_id else None
    dept_name = dept_obj.name if dept_obj else ""

    _, num_days = calendar.monthrange(year, month_num)
    days = [date(year, month_num, 1) + timedelta(days=i) for i in range(num_days)]

    employees = {e.id: _emp_to_dict(e) for e in Employee.query.all()}
    start_str = f"{year:04d}-{month_num:02d}-01"
    end_str = f"{year:04d}-{month_num:02d}-{num_days:02d}"
    _att_recs = AttendanceRecord.query.filter(
        AttendanceRecord.date >= start_str,
        AttendanceRecord.date <= end_str,
    ).all()
    attendance = {}
    for r in _att_recs:
        attendance.setdefault(r.date, {})[r.emp_id] = {
            "check_in": r.check_in_time,
            "check_out": r.check_out_time,
        }
    _ov_recs = TimesheetOverride.query.all()
    overrides = {}
    for r in _ov_recs:
        overrides.setdefault(r.emp_id, {})[r.date] = r.symbol
    holidays_set = get_holidays_set(year)
    missing_holiday_year = is_holiday_year_missing(year)

    scoped_employees = {eid: e for eid, e in employees.items() if e.get("dept_id") == dept_id} if dept_id else {}

    grid_rows = []
    for emp_id_key, emp in scoped_employees.items():
        schedule = emp.get("schedule", {"start": "09:00", "end": "18:00", "work_days": [1, 2, 3, 4, 5]})
        cells = []
        for d in days:
            sym = compute_symbol(d, emp_id_key, attendance, overrides, schedule, holidays_set)
            auto = compute_symbol(d, emp_id_key, attendance, {}, schedule, holidays_set)
            cells.append({"sym": sym, "auto": auto, "date": d.isoformat()})
        totals = compute_employee_totals([c["sym"] for c in cells], schedule)
        grid_rows.append((emp_id_key, emp.get("name", emp_id_key), cells, totals))

    return render_template(
        "timesheet_partial.html",
        username=username,
        role="dept_admin",
        dept_id=dept_id or "",
        dept_name=dept_name,
        dept_options=[],
        month_str=month_str,
        year=year,
        month_num=month_num,
        month_name=MONTHS_RU.get(month_num, ""),
        days=days,
        grid_rows=grid_rows,
        holidays_set=holidays_set,
        missing_holiday_year=missing_holiday_year,
        can_edit=True,
        partial_url="/dept_admin/partial/timesheet",
        content_el_id="inlineDeptTimesheetContent",
    )


# ─── API: Orgs ────────────────────────────────────────────────────────────────

@app.route("/api/orgs", methods=["GET"])
@require_role("superadmin", "org_admin", "dept_admin")
def list_orgs():
    return jsonify([_org_to_dict(o) for o in Organization.query.all()])


@app.route("/api/orgs", methods=["POST"])
@require_role("superadmin")
def create_org():
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Название организации не может быть пустым"}), 400
    org_id = str(uuid.uuid4())
    # Build a seen-set of all existing tokens to guarantee uniqueness
    seen_tokens = set()
    for existing in Organization.query.all():
        if existing.org_token:
            seen_tokens.add(existing.org_token)
        if existing.reg_token:
            seen_tokens.add(existing.reg_token)
    org_token = generate_unique_token(seen_tokens)
    seen_tokens.add(org_token)
    reg_token = generate_unique_token(seen_tokens)
    try:
        db.session.add(Organization(
            id=org_id,
            name=name,
            description=data.get("description", ""),
            created_at=datetime.now().isoformat(),
            org_token=org_token,
            reg_token=reg_token,
            kiosk_pin=hash_pin("0000"),
            reg_pin=hash_pin("1234"),
            reg_token_expires=None,
            kiosk_display_name=name,
        ))
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    write_audit("org_create", target_type="org", target_id=org_id,
                new_value={"name": name})
    return jsonify({"id": org_id, "status": "created"})


@app.route("/api/orgs/<org_id>", methods=["PUT"])
@require_role("superadmin")
def update_org(org_id):
    org = Organization.query.get(org_id)
    if not org:
        return jsonify({"error": "Организация не найдена"}), 404
    data = request.json or {}
    old_name = org.name
    old_description = org.description
    if "name" in data:
        name = data["name"].strip()
        if not name:
            return jsonify({"error": "Название организации не может быть пустым"}), 400
        org.name = name
    if "description" in data:
        org.description = data["description"]
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    write_audit("org_update", target_type="org", target_id=org_id,
                old_value={"name": old_name, "description": old_description},
                new_value={k: data[k] for k in ("name", "description") if k in data})
    return jsonify({"status": "updated"})


@app.route("/api/orgs/<org_id>", methods=["DELETE"])
@require_role("superadmin")
def delete_org(org_id):
    org = Organization.query.get(org_id)
    if not org:
        return jsonify({"error": "Организация не найдена"}), 404
    if Employee.query.filter_by(org_id=org_id).count() > 0:
        return jsonify({"error": "Организация содержит сотрудников"}), 409
    old_name = org.name
    try:
        db.session.delete(org)
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    write_audit("org_delete", target_type="org", target_id=org_id,
                old_value={"name": old_name})
    return jsonify({"status": "deleted"})


@app.route("/api/orgs/<org_id>/settings", methods=["PATCH"])
@require_role("superadmin", "org_admin")
def update_org_settings(org_id):
    org = Organization.query.get(org_id)
    if not org:
        return jsonify({"error": "Организация не найдена"}), 404
    caller_role = session.get("role")
    if caller_role == "org_admin" and session.get("org_id") != org_id:
        return jsonify({"error": "forbidden"}), 403
    data = request.json or {}

    # kiosk_pin — 4 digits; stored as bcrypt hash; used only for device registration
    if "kiosk_pin" in data:
        pin = data["kiosk_pin"]
        if pin:
            if len(str(pin)) != 4 or not str(pin).isdigit():
                return jsonify({"error": "PIN должен быть 4-значным числом"}), 400
            org.kiosk_pin = hash_pin(pin)
        else:
            org.kiosk_pin = None

    # reg_pin — validate 4 digits; store as bcrypt hash
    if "reg_pin" in data:
        pin = data["reg_pin"]
        if pin:
            if len(str(pin)) != 4 or not str(pin).isdigit():
                return jsonify({"error": "PIN должен быть 4-значным числом"}), 400
            org.reg_pin = hash_pin(pin)
        else:
            org.reg_pin = None

    # regen_reg_token — generate a new unique 8-hex reg_token
    if data.get("regen_reg_token"):
        seen = set()
        for existing in Organization.query.all():
            if existing.org_token:
                seen.add(existing.org_token)
            if existing.reg_token:
                seen.add(existing.reg_token)
        org.reg_token = generate_unique_token(seen)

    # reg_token_expires — store ISO string or clear to None
    if "reg_token_expires" in data:
        expires = data["reg_token_expires"]
        if expires:
            try:
                datetime.fromisoformat(expires)
            except (ValueError, TypeError):
                return jsonify({"error": "Неверный формат даты (ожидается ISO 8601)"}), 400
            org.reg_token_expires = expires
        else:
            org.reg_token_expires = None

    # kiosk_display_name — store trimmed string (allow empty)
    if "kiosk_display_name" in data:
        org.kiosk_display_name = str(data["kiosk_display_name"]).strip()

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    return jsonify({"status": "updated", "reg_token": org.reg_token})


# ─── API: Kiosk device registration & management ──────────────────────────────

@app.route("/api/kiosk/<org_token>/register_device", methods=["POST"])
def register_kiosk_device(org_token):
    """Validate kiosk PIN and register a new device; sets httpOnly cookie in response."""
    org = Organization.query.filter_by(org_token=org_token).first()
    if not org:
        return jsonify({"error": "not_found"}), 404

    data = request.json or {}
    pin = str(data.get("pin", ""))
    device_name = str(data.get("device_name", "Киоск")).strip() or "Киоск"

    stored = org.kiosk_pin
    if not stored:
        return jsonify({"error": "no_pin_set",
                        "message": "PIN киоска не настроен — обратитесь к администратору"}), 400
    if not bcrypt.checkpw(pin.encode(), stored.encode()):
        return jsonify({"error": "wrong_pin"}), 401

    raw_token = secrets.token_urlsafe(32)
    token_hash = _hash_device_token(raw_token)
    device = KioskDevice(
        id=str(uuid.uuid4()),
        org_id=org.id,
        device_token_hash=token_hash,
        device_name=device_name,
        created_at=datetime.now().isoformat(),
        last_seen_at=None,
    )
    db.session.add(device)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500

    resp = jsonify({"status": "registered", "device_name": device_name})
    resp.set_cookie(
        _device_cookie_name(org_token),
        raw_token,
        max_age=365 * 24 * 3600,
        httponly=True,
        samesite="Lax",
        secure=False,   # set True when enforcing HTTPS at Flask level
    )
    print(f"DEVICE_REGISTERED: org_id={org.id!r} name={device_name!r} id={device.id!r}", flush=True)
    return resp


@app.route("/org/devices")
@require_role("org_admin")
def org_devices_page():
    """Org-admin device management page."""
    caller_org_id = session.get("org_id")
    org = Organization.query.get(caller_org_id)
    if not org:
        return redirect(url_for("org_admin_page"))
    devices = (KioskDevice.query
               .filter_by(org_id=caller_org_id)
               .order_by(KioskDevice.created_at.desc())
               .all())
    return render_template("devices.html",
                           org_name=org.name,
                           org_token=org.org_token or "",
                           devices=devices,
                           has_kiosk_pin=bool(org.kiosk_pin))


@app.route("/api/kiosk/<org_token>/devices", methods=["GET"])
@require_role("org_admin", "superadmin")
def list_kiosk_devices(org_token):
    org = Organization.query.filter_by(org_token=org_token).first()
    if not org:
        return jsonify({"error": "not_found"}), 404
    if session.get("role") == "org_admin" and session.get("org_id") != org.id:
        return jsonify({"error": "forbidden"}), 403
    devices = (KioskDevice.query.filter_by(org_id=org.id)
               .order_by(KioskDevice.created_at.desc()).all())
    return jsonify([{
        "id": d.id, "device_name": d.device_name,
        "created_at": d.created_at, "last_seen_at": d.last_seen_at,
    } for d in devices])


@app.route("/api/kiosk/<org_token>/devices/<device_id>", methods=["DELETE"])
@require_role("org_admin", "superadmin")
def revoke_kiosk_device(org_token, device_id):
    org = Organization.query.filter_by(org_token=org_token).first()
    if not org:
        return jsonify({"error": "not_found"}), 404
    if session.get("role") == "org_admin" and session.get("org_id") != org.id:
        return jsonify({"error": "forbidden"}), 403
    device = KioskDevice.query.get(device_id)
    if not device or device.org_id != org.id:
        return jsonify({"error": "not_found"}), 404
    db.session.delete(device)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    return jsonify({"status": "revoked"})


@app.route("/api/kiosk/<org_token>/devices/<device_id>", methods=["PATCH"])
@require_role("org_admin", "superadmin")
def rename_kiosk_device(org_token, device_id):
    org = Organization.query.filter_by(org_token=org_token).first()
    if not org:
        return jsonify({"error": "not_found"}), 404
    if session.get("role") == "org_admin" and session.get("org_id") != org.id:
        return jsonify({"error": "forbidden"}), 403
    device = KioskDevice.query.get(device_id)
    if not device or device.org_id != org.id:
        return jsonify({"error": "not_found"}), 404
    data = request.json or {}
    name = str(data.get("device_name", "")).strip()
    if not name:
        return jsonify({"error": "Название не может быть пустым"}), 400
    device.device_name = name
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    return jsonify({"status": "renamed"})


# ─── API: Depts ───────────────────────────────────────────────────────────────

@app.route("/api/depts", methods=["GET"])
@require_role("superadmin", "org_admin", "hr_viewer", "dept_admin")
def list_depts():
    caller_role = session.get("role")
    caller_org_id = session.get("org_id")
    if caller_role in ("org_admin", "hr_viewer", "dept_admin"):
        result = [_dept_to_dict(d) for d in Department.query.filter_by(org_id=caller_org_id).all()]
    else:
        result = [_dept_to_dict(d) for d in Department.query.all()]
    return jsonify(result)


@app.route("/api/depts", methods=["POST"])
@require_role("superadmin", "org_admin")
def create_dept():
    caller_role = session.get("role")
    caller_org_id = session.get("org_id")
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Название отдела не может быть пустым"}), 400
    # org_admin always creates within their own org regardless of request body
    if caller_role == "org_admin":
        target_org_id = caller_org_id
    else:
        target_org_id = data.get("org_id")
    dept_id = str(uuid.uuid4())
    try:
        db.session.add(Department(
            id=dept_id,
            org_id=target_org_id,
            name=name,
            head_name=data.get("head_name", ""),
            created_at=datetime.now().isoformat(),
        ))
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    write_audit("dept_create", target_type="dept", target_id=dept_id,
                new_value={"name": name, "org_id": target_org_id})
    return jsonify({"id": dept_id, "status": "created"})


@app.route("/api/depts/<dept_id>", methods=["PUT"])
@require_role("superadmin", "org_admin")
def update_dept(dept_id):
    caller_role = session.get("role")
    caller_org_id = session.get("org_id")
    dept = Department.query.get(dept_id)
    if not dept:
        return jsonify({"error": "Отдел не найден"}), 404
    if caller_role == "org_admin" and dept.org_id != caller_org_id:
        return jsonify({"error": "forbidden"}), 403
    data = request.json or {}
    old_name = dept.name
    old_head_name = dept.head_name
    if "name" in data:
        name = data["name"].strip()
        if not name:
            return jsonify({"error": "Название отдела не может быть пустым"}), 400
        dept.name = name
    if "head_name" in data:
        dept.head_name = data["head_name"]
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    write_audit("dept_update", target_type="dept", target_id=dept_id,
                old_value={"name": old_name, "head_name": old_head_name},
                new_value={k: data[k] for k in ("name", "head_name") if k in data})
    return jsonify({"status": "updated"})


@app.route("/api/depts/<dept_id>", methods=["DELETE"])
@require_role("superadmin", "org_admin")
def delete_dept(dept_id):
    caller_role = session.get("role")
    caller_org_id = session.get("org_id")
    dept = Department.query.get(dept_id)
    if not dept:
        return jsonify({"error": "Отдел не найден"}), 404
    if caller_role == "org_admin" and dept.org_id != caller_org_id:
        return jsonify({"error": "forbidden"}), 403
    if Employee.query.filter_by(dept_id=dept_id).count() > 0:
        return jsonify({"error": "Отдел содержит сотрудников"}), 409
    old_dept_name = dept.name
    try:
        db.session.delete(dept)
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    write_audit("dept_delete", target_type="dept", target_id=dept_id,
                old_value={"name": old_dept_name})
    return jsonify({"status": "deleted"})


# ─── API: Employees ───────────────────────────────────────────────────────────

@app.route("/api/employees", methods=["GET"])
@require_role("superadmin", "org_admin", "hr_viewer", "dept_admin")
def get_employees():
    role = session.get("role")
    org_id = session.get("org_id")
    dept_id = session.get("dept_id")
    if role == "superadmin":
        emps = Employee.query.all()
    elif role in ("org_admin", "hr_viewer") and org_id:
        emps = Employee.query.filter_by(org_id=org_id).all()
    elif role == "dept_admin" and dept_id:
        emps = Employee.query.filter_by(dept_id=dept_id).all()
    else:
        emps = Employee.query.all()
    return jsonify({e.id: _emp_to_dict(e) for e in emps})

@app.route("/api/employees", methods=["POST"])
@require_role("superadmin", "org_admin", "dept_admin")
def add_employee():
    caller_role = session.get("role")
    caller_dept_id = session.get("dept_id")
    caller_org_id = session.get("org_id")
    data = request.json or {}

    # Scope gate: dept_admin may only create in their own dept.
    # Allow omitting dept_id (defaults to caller's dept); reject only if explicitly different.
    target_dept_id = data.get("dept_id")
    if caller_role == "dept_admin" and target_dept_id and target_dept_id != caller_dept_id:
        return jsonify({"error": "forbidden"}), 403

    emp_id = str(int(time.time() * 1000))
    label = Employee.query.count() + 1

    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "ФИО обязательно"}), 400

    # Determine org_id/dept_id: dept_admin defaults to session values when omitted
    org_id = data.get("org_id") if caller_role != "dept_admin" else (data.get("org_id") or caller_org_id)
    dept_id = target_dept_id if caller_role != "dept_admin" else caller_dept_id

    sched = data.get("schedule", {"start": "09:00", "end": "18:00", "work_days": [1, 2, 3, 4, 5]})
    try:
        db.session.add(Employee(
            id=emp_id,
            name=name,
            role=data.get("role", "employee"),
            label=label,
            registered_at=datetime.now().isoformat(),
            face_count=0,
            org_id=org_id,
            dept_id=dept_id,
        ))
        db.session.add(EmployeeSchedule(
            emp_id=emp_id,
            start_time=sched.get("start", "09:00"),
            end_time=sched.get("end", "18:00"),
            work_days_json=json.dumps(sched.get("work_days", [1, 2, 3, 4, 5])),
        ))
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    os.makedirs(os.path.join(FACES_DIR, emp_id), exist_ok=True)
    write_audit("employee_register", target_type="employee", target_id=emp_id,
                new_value={"name": name, "org_id": org_id, "dept_id": dept_id})
    return jsonify({"id": emp_id, "status": "created"})


@app.route("/api/employees/<emp_id>", methods=["PATCH"])
@require_role("superadmin", "org_admin")
def update_employee_assignment(emp_id):
    """ORG-04: reassign employee to another dept; org_admin restricted to own org scope."""
    caller_role = session.get("role")
    caller_org_id = session.get("org_id")
    emp = Employee.query.get(emp_id)
    if not emp:
        return jsonify({"error": "Сотрудник не найден"}), 404

    # Scope gate: org_admin can only edit employees in their own org
    if caller_role == "org_admin" and emp.org_id != caller_org_id:
        return jsonify({"error": "forbidden"}), 403

    data = request.json or {}

    # Whitelist: dept_id/org_id for superadmin; dept_id/name/role/iin for org_admin
    allowed_keys = {"dept_id", "org_id", "name", "role", "iin"} if caller_role == "superadmin" else {"dept_id", "name", "role", "iin"}
    update_data = {k: v for k, v in data.items() if k in allowed_keys}

    if "dept_id" in update_data:
        target_dept_id = update_data["dept_id"]
        if caller_role == "org_admin":
            # Verify target dept belongs to caller's org
            target_dept = Department.query.get(target_dept_id)
            if not target_dept or target_dept.org_id != caller_org_id:
                return jsonify({"error": "forbidden"}), 403

    if "dept_id" in update_data:
        emp.dept_id = update_data["dept_id"]
    if "org_id" in update_data:
        emp.org_id = update_data["org_id"]
    if "name" in update_data:
        name = update_data["name"].strip()
        if not name:
            return jsonify({"error": "ФИО обязательно"}), 400
        emp.name = name
    if "role" in update_data:
        emp.role = update_data["role"]
    if "iin" in update_data:
        emp.iin = (update_data["iin"] or "").strip() or None

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    return jsonify({"status": "updated", "employee": _emp_to_dict(emp)})

@app.route("/api/employees/<emp_id>", methods=["DELETE"])
@require_role("superadmin", "org_admin", "dept_admin")
def delete_employee(emp_id):
    emp = Employee.query.get(emp_id)
    if not emp:
        return jsonify({"status": "deleted"})
    role = session.get("role")
    if role == "dept_admin" and emp.dept_id != session.get("dept_id"):
        return jsonify({"error": "forbidden"}), 403
    if role == "org_admin" and emp.org_id != session.get("org_id"):
        return jsonify({"error": "forbidden"}), 403
    try:
        # Remove associated schedule first
        sched = EmployeeSchedule.query.filter_by(emp_id=emp_id).first()
        if sched:
            db.session.delete(sched)
        db.session.delete(emp)
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    emp_dir = os.path.join(FACES_DIR, emp_id)
    if os.path.exists(emp_dir):
        shutil.rmtree(emp_dir)
    train_recognizer()
    return jsonify({"status": "deleted"})

@app.route("/api/employees/<emp_id>", methods=["GET"])
@require_role("superadmin", "org_admin", "dept_admin")
def get_employee(emp_id):
    emp = Employee.query.get(emp_id)
    if not emp:
        return jsonify({"error": "Сотрудник не найден"}), 404
    return jsonify(_emp_to_dict(emp))

@app.route("/api/employees/<emp_id>/schedule", methods=["PATCH"])
@require_role("superadmin", "org_admin", "dept_admin")
def update_employee_schedule(emp_id):
    """T13-06: Update per-employee work schedule with HH:MM validation."""
    caller_role = session.get("role")
    caller_dept_id = session.get("dept_id")
    emp = Employee.query.get(emp_id)
    if not emp:
        return jsonify({"error": "Сотрудник не найден"}), 404

    # Scope gate: dept_admin may only edit employees in their own dept
    if caller_role == "dept_admin" and emp.dept_id != caller_dept_id:
        return jsonify({"error": "forbidden"}), 403

    data = request.json or {}
    start = data.get("start", "")
    end = data.get("end", "")
    work_days = data.get("work_days")

    # Validate HH:MM format with range check
    def valid_time(t):
        import re
        if not re.match(r'^\d{2}:\d{2}$', str(t)):
            return False
        h, m = map(int, t.split(":"))
        return 0 <= h <= 23 and 0 <= m <= 59

    if not valid_time(start) or not valid_time(end):
        return jsonify({"error": "Неверный формат времени"}), 400

    # Validate work_days is a list of ints in 1..7
    if not isinstance(work_days, list) or not work_days:
        return jsonify({"error": "Неверный формат рабочих дней"}), 400
    for d in work_days:
        if not isinstance(d, int) or d < 1 or d > 7:
            return jsonify({"error": "Неверный формат рабочих дней"}), 400

    # Upsert EmployeeSchedule row
    sched = EmployeeSchedule.query.filter_by(emp_id=emp_id).first()
    if sched:
        sched.start_time = start
        sched.end_time = end
        sched.work_days_json = json.dumps(work_days)
    else:
        db.session.add(EmployeeSchedule(
            emp_id=emp_id,
            start_time=start,
            end_time=end,
            work_days_json=json.dumps(work_days),
        ))
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    return jsonify({"status": "updated"})

@app.route("/api/employees/<emp_id>/reset", methods=["POST"])
@require_role("superadmin", "org_admin", "dept_admin")
def reset_employee_face(emp_id):
    emp = Employee.query.get(emp_id)
    if not emp:
        return jsonify({"error": "Сотрудник не найден"}), 404
    role = session.get("role")
    if role == "dept_admin" and emp.dept_id != session.get("dept_id"):
        return jsonify({"error": "forbidden"}), 403
    if role == "org_admin" and emp.org_id != session.get("org_id"):
        return jsonify({"error": "forbidden"}), 403
    emp_dir = os.path.join(FACES_DIR, emp_id)
    if os.path.exists(emp_dir):
        shutil.rmtree(emp_dir)
    os.makedirs(emp_dir, exist_ok=True)
    emp.face_count = 0
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    train_recognizer()
    return jsonify({"status": "reset", "face_count": 0})


@app.route("/api/employees/import_xlsx", methods=["POST"])
@require_role("superadmin", "org_admin")
def import_employees_xlsx():
    """Import employees from uploaded Excel file (medical staff report format).

    Expected format: header row at row 6, data from row 7.
    Columns: Фамилия(4), Имя(5), Отчество(6), ИИН(7), Отделение(9).
    Matches employees by IIN; creates if not found, updates name/dept otherwise.
    """
    if "file" not in request.files:
        return jsonify({"error": "Файл не передан"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Поддерживаются только файлы .xlsx"}), 400

    caller_role = session.get("role")
    caller_org_id = session.get("org_id")

    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
    except Exception:
        return jsonify({"error": "Не удалось прочитать файл Excel"}), 400

    # Determine next label for new employees
    max_label_row = db.session.execute(
        db.select(db.func.max(Employee.label))
    ).scalar()
    next_label = (max_label_row or 0) + 1

    created = updated = skipped = 0
    errors = []

    # Collect all depts for name→id matching
    dept_by_name = {d.name.lower(): d for d in Department.query.all()}

    # Find data rows: skip until we hit the header row (contains "ИИН"), then read data
    data_start = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and any(isinstance(c, str) and "ИИН" in c for c in row):
            data_start = i + 1
            break

    if data_start is None:
        return jsonify({"error": "Не найдена строка заголовка с полем ИИН"}), 400

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not any(row):
            continue
        try:
            last_name = str(row[4] or "").strip()
            first_name = str(row[5] or "").strip()
            middle_name = str(row[6] or "").strip()
            iin = str(row[7] or "").strip()
            dept_name_raw = str(row[9] or "").strip()
        except IndexError:
            skipped += 1
            continue

        if not last_name and not first_name:
            skipped += 1
            continue

        full_name = " ".join(p for p in [last_name, first_name, middle_name] if p)
        if not full_name:
            skipped += 1
            continue

        # Resolve dept
        dept_id = None
        if dept_name_raw:
            dept_obj = dept_by_name.get(dept_name_raw.lower())
            if dept_obj:
                if caller_role == "org_admin" and dept_obj.org_id != caller_org_id:
                    dept_id = None
                else:
                    dept_id = dept_obj.id

        # Try to find existing employee by IIN
        existing = None
        if iin:
            existing = Employee.query.filter_by(iin=iin).first()

        if existing:
            existing.name = full_name
            if dept_id:
                existing.dept_id = dept_id
            if iin:
                existing.iin = iin
            updated += 1
        else:
            emp_id = str(uuid.uuid4())
            new_emp = Employee(
                id=emp_id,
                name=full_name,
                role="employee",
                label=next_label,
                face_count=0,
                registered_at=None,
                org_id=caller_org_id,
                dept_id=dept_id,
                iin=iin or None,
            )
            db.session.add(new_emp)
            next_label += 1
            created += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Ошибка при сохранении: {e}"}), 500

    return jsonify({"created": created, "updated": updated, "skipped": skipped, "errors": errors})


# ─── API: Dashboards ──────────────────────────────────────────────────────────

@app.route("/api/superadmin_stats", methods=["GET"])
@require_role("superadmin")
def superadmin_stats():
    """DASH-01: System-wide stats for superadmin dashboard."""
    today = date.today().isoformat()
    checkins_today = AttendanceRecord.query.filter(
        AttendanceRecord.date == today,
        AttendanceRecord.check_in_time != None,  # noqa: E711
    ).count()
    return jsonify({
        "orgs": Organization.query.count(),
        "employees": Employee.query.count(),
        "checkins_today": checkins_today,
    })


@app.route("/api/dept_attendance_today", methods=["GET"])
@require_role("dept_admin", "org_admin", "hr_viewer", "superadmin")
def dept_attendance_today():
    """DASH-02: Today's attendance scoped by caller role (dept_admin→dept, org_admin→org, superadmin→all)."""
    role = session.get("role")
    dept_id = session.get("dept_id")
    org_id = session.get("org_id")

    today = date.today().isoformat()
    today_weekday = date.today().weekday() + 1  # ISO 1=Mon, 7=Sun

    # Build attendance dict adapter for today only
    if role == "dept_admin":
        emps = Employee.query.filter_by(dept_id=dept_id).all()
    elif role in ("org_admin", "hr_viewer"):
        emps = Employee.query.filter_by(org_id=org_id).all()
    else:  # superadmin — all employees
        emps = Employee.query.all()
    scoped = {e.id: _emp_to_dict(e) for e in emps}

    # Query today's attendance records for scoped employees only
    scoped_ids = list(scoped.keys())
    _att_recs = AttendanceRecord.query.filter(
        AttendanceRecord.date == today,
        AttendanceRecord.emp_id.in_(scoped_ids),
    ).all() if scoped_ids else []
    today_records = {r.emp_id: {"check_in": r.check_in_time, "check_out": r.check_out_time}
                     for r in _att_recs}

    result = []
    present = absent = late = 0

    for eid, emp in scoped.items():
        schedule = emp.get("schedule", {"start": "09:00", "end": "18:00", "work_days": [1, 2, 3, 4, 5]})
        work_days = schedule.get("work_days", [1, 2, 3, 4, 5])

        if today_weekday not in work_days:
            continue  # Day off — do not count as absent (A3)

        rec = today_records.get(eid)
        check_in = rec.get("check_in") if rec else None
        check_out = rec.get("check_out") if rec else None

        # Late detection: check_in > schedule.start + 15 min grace (A1)
        schedule_start = schedule.get("start", "09:00")
        sh, sm = map(int, schedule_start.split(":"))
        late_m = sm + 15
        if late_m < 60:
            late_threshold = f"{sh:02d}:{late_m:02d}:00"
        else:
            late_threshold = f"{sh + 1:02d}:{late_m % 60:02d}:00"

        if check_in:
            if check_in > late_threshold:
                status = "late"
                late += 1
            else:
                status = "present"
                present += 1
        else:
            status = "absent"
            absent += 1

        result.append({
            "emp_id": eid,
            "name": emp["name"],
            "check_in": check_in,
            "check_out": check_out,
            "status": status,
            "schedule": f"{schedule.get('start')} – {schedule.get('end')}",
            "face_count": emp.get("face_count", 0),
        })

    return jsonify({
        "employees": result,
        "stats": {"present": present, "absent": absent, "late": late},
    })

# ─── API: Face Registration ───────────────────────────────────────────────────

@app.route("/api/register_face", methods=["POST"])
@require_role("superadmin", "org_admin", "dept_admin")
def register_face():
    data = request.get_json(silent=True) or {}
    if not data.get("emp_id") or not data.get("image"):
        return jsonify({"error": "emp_id and image required"}), 400
    emp_id = data["emp_id"]
    emp = Employee.query.get(emp_id)
    if not emp:
        return jsonify({"error": "Сотрудник не найден"}), 404

    img = decode_image(data["image"])
    face_roi, bbox = extract_face(img)
    if face_roi is None:
        return jsonify({"error": "Лицо не обнаружено. Убедитесь, что лицо хорошо видно в кадре."}), 400

    emp_dir = os.path.join(FACES_DIR, emp_id)
    os.makedirs(emp_dir, exist_ok=True)
    count = emp.face_count + 1
    cv2.imwrite(os.path.join(emp_dir, f"face_{count}.jpg"), face_roi)
    emp.face_count = count
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500
    train_recognizer()
    return jsonify({"status": "saved", "count": count, "bbox": bbox})

@app.route("/api/detect", methods=["POST"])
def detect_face_only():
    """Detect face bbox without saving — used for live preview overlay."""
    try:
        data = request.json
        img = decode_image(data["image"])
        if img is None:
            return jsonify({"face": False})
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(80, 80))
        if len(faces) == 0:
            return jsonify({"face": False})
        x, y, w, h = max(faces, key=lambda f: f[2]*f[3])
        faces_strict = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=9, minSize=(80, 80))
        quality = "good" if len(faces_strict) > 0 else "ok"
        return jsonify({"face": True, "bbox": [int(x), int(y), int(w), int(h)], "quality": quality})
    except Exception:
        return jsonify({"face": False})

# ─── API: Recognition ─────────────────────────────────────────────────────────

@app.route("/api/recognize", methods=["POST"])
def recognize():
    global recognizer_trained
    if Employee.query.count() == 0:
        return jsonify({"error": "no_employees"}), 400

    if not recognizer_trained:
        ok = train_recognizer()
        if not ok:
            return jsonify({"error": "not_trained"}), 400

    data = request.get_json(silent=True) or {}
    if "image" not in data:
        return jsonify({"error": "image required"}), 400
    img = decode_image(data["image"])
    face_roi, bbox = extract_face(img)
    if face_roi is None:
        return jsonify({"error": "no_face"}), 400

    label, confidence = recognizer.predict(face_roi)
    # LBPH: lower confidence = better. Threshold ~80. Convert to % (0–100 good to bad).
    conf_pct = max(0, min(100, round(100 - (confidence / 80 * 100))))

    if confidence > 80:
        append_log({"ts": datetime.now().isoformat(), "event": "unknown",
                    "confidence_raw": float(confidence), "confidence_pct": conf_pct})
        return jsonify({"error": "unknown", "confidence": float(confidence)}), 400

    emp = Employee.query.filter_by(label=label).first()
    if not emp:
        return jsonify({"error": "unknown"}), 400

    # Org filter: if request specifies org_id, reject matches from other orgs
    req_org_id = data.get("org_id")
    if req_org_id and emp.org_id != req_org_id:
        return jsonify({"error": "unknown"}), 400

    # Device auth: org-linked kiosks must present a valid registered device cookie
    if req_org_id:
        req_org = Organization.query.get(req_org_id)
        if req_org and req_org.org_token:
            device = _check_device_auth(req_org)
            if not device:
                return jsonify({"error": "device_unauthorized"}), 403
            _maybe_update_last_seen(device)

    dept_name = None
    if emp.dept_id:
        dept = Department.query.get(emp.dept_id)
        if dept:
            dept_name = dept.name

    today = date.today().isoformat()
    emp_id = emp.id
    emp_dict = _emp_to_dict(emp)
    now_dt = datetime.now()
    now = now_dt.strftime("%H:%M:%S")
    schedule = emp_dict.get("schedule", {})
    start = schedule.get("start", "09:00")
    is_late = now > _time_threshold(start, 15)

    # Mode-aware attendance logic (kiosk «Прибыл»/«Убыл» buttons).
    # If mode is absent or invalid, fall back to legacy state machine for backward compat.
    mode = data.get("mode")
    rec = AttendanceRecord.query.filter_by(emp_id=emp_id, date=today).first()
    try:
        if mode == "arrive":
            if rec is None:
                rec = AttendanceRecord(
                    emp_id=emp_id,
                    date=today,
                    check_in_time=now,
                    check_out_time=None,
                    event_type="check_in",
                )
                db.session.add(rec)
                event = "check_in"
            elif rec.check_in_time:
                # Arrival already recorded — do NOT overwrite
                event = "already_in"
            else:
                rec.check_in_time = now
                rec.event_type = "check_in"
                event = "check_in"
        elif mode == "depart":
            if rec is None or rec.check_in_time is None:
                # No arrival recorded — refuse departure
                return jsonify({"error": "no_checkin",
                                "message": "Сначала отметьте приход"}), 409
            if rec.check_out_time:
                # Departure already recorded — do NOT overwrite
                event = "already_out"
            else:
                rec.check_out_time = now
                rec.event_type = "check_out"
                event = "check_out"
        else:
            # Legacy state machine (no mode parameter — backward compatible)
            if rec is None:
                rec = AttendanceRecord(
                    emp_id=emp_id,
                    date=today,
                    check_in_time=now,
                    check_out_time=None,
                    event_type="check_in",
                )
                db.session.add(rec)
                event = "check_in"
            elif rec.check_out_time is None:
                rec.check_out_time = now
                rec.event_type = "check_out"
                event = "check_out"
            else:
                event = "already_done"
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500

    record_dict = {"check_in": rec.check_in_time, "check_out": rec.check_out_time}
    append_log({"ts": now_dt.isoformat(), "emp_id": emp_id, "name": emp.name,
                "event": event, "confidence_raw": float(confidence), "confidence_pct": conf_pct})

    return jsonify({
        "status": "ok",
        "employee": emp_dict,
        "event": event,
        "record": record_dict,
        "confidence": float(confidence),
        "confidence_pct": conf_pct,
        "is_late": is_late and event == "check_in",
        "bbox": bbox,
        "dept_name": dept_name,
    })

# ─── API: Kiosk public log ────────────────────────────────────────────────────

@app.route("/api/kiosk_log")
def kiosk_log():
    """Public endpoint for kiosk attendance log — no auth, today only, optional org filter."""
    org_id = request.args.get("org_id")
    today = date.today().isoformat()
    if org_id:
        all_emps = Employee.query.filter_by(org_id=org_id).all()
    else:
        all_emps = Employee.query.all()
    emp_ids = [e.id for e in all_emps]
    emp_by_id = {e.id: e for e in all_emps}
    _att_recs = AttendanceRecord.query.filter(
        AttendanceRecord.date == today,
        AttendanceRecord.emp_id.in_(emp_ids),
        AttendanceRecord.check_in_time != None,  # noqa: E711
    ).all() if emp_ids else []
    result = []
    for r in _att_recs:
        emp = emp_by_id.get(r.emp_id)
        if emp:
            result.append({"name": emp.name, "role": emp.role,
                           "check_in": r.check_in_time, "check_out": r.check_out_time})
    return jsonify(result)

# ─── API: Attendance ──────────────────────────────────────────────────────────

@app.route("/api/attendance", methods=["GET"])
@require_role("superadmin", "org_admin", "hr_viewer", "dept_admin")
def get_attendance():
    day = request.args.get("date", date.today().isoformat())
    role = session.get("role")
    org_id = session.get("org_id")
    dept_id = session.get("dept_id")
    if role == "superadmin":
        emps = Employee.query.all()
    elif role in ("org_admin", "hr_viewer") and org_id:
        emps = Employee.query.filter_by(org_id=org_id).all()
    elif role == "dept_admin" and dept_id:
        emps = Employee.query.filter_by(dept_id=dept_id).all()
    else:
        emps = Employee.query.all()
    employees = {e.id: _emp_to_dict(e) for e in emps}
    # Build attendance dict for this day scoped to visible employees
    emp_ids = list(employees.keys())
    _att_recs = AttendanceRecord.query.filter(
        AttendanceRecord.date == day,
        AttendanceRecord.emp_id.in_(emp_ids)
    ).all()
    day_data = {r.emp_id: {"check_in": r.check_in_time, "check_out": r.check_out_time}
                for r in _att_recs}
    result = []
    for emp_id, emp in employees.items():
        rec = day_data.get(emp_id, {})
        check_in = rec.get("check_in")
        check_out = rec.get("check_out")
        duration = None
        duration_minutes = None
        if check_in and check_out:
            fmt = "%H:%M:%S"
            delta = datetime.strptime(check_out, fmt) - datetime.strptime(check_in, fmt)
            total = int(delta.total_seconds())
            duration_minutes = total // 60
            duration = f"{total//3600}ч {(total%3600)//60}мин"
        sched = emp.get("schedule", {})
        result.append({
            "emp_id": emp_id,
            "name": emp["name"],
            "role": emp["role"],
            "check_in": check_in,
            "check_out": check_out,
            "duration": duration,
            "duration_minutes": duration_minutes,
            "schedule_start": (sched.get("start") or "09:00") + ":00",
            "schedule_end": (sched.get("end") or "18:00") + ":00",
        })
    return jsonify(result)

@app.route("/api/attendance/dates", methods=["GET"])
@require_role("superadmin", "org_admin", "hr_viewer", "dept_admin")
def get_dates():
    dates = db.session.execute(
        db.select(AttendanceRecord.date).distinct().order_by(AttendanceRecord.date.desc())
    ).scalars().all()
    return jsonify(list(dates))

@app.route("/api/stats", methods=["GET"])
@require_role("superadmin", "org_admin", "hr_viewer", "dept_admin")
def get_stats():
    from_date = request.args.get("from")
    to_date = request.args.get("to")
    role = session.get("role")
    org_id = session.get("org_id")
    dept_id = session.get("dept_id")
    if role == "superadmin":
        emps = Employee.query.all()
    elif role in ("org_admin", "hr_viewer") and org_id:
        emps = Employee.query.filter_by(org_id=org_id).all()
    elif role == "dept_admin" and dept_id:
        emps = Employee.query.filter_by(dept_id=dept_id).all()
    else:
        emps = Employee.query.all()
    employees = {e.id: _emp_to_dict(e) for e in emps}

    # Query attendance records scoped to visible employees with optional date filters
    emp_ids = list(employees.keys())
    q = AttendanceRecord.query.filter(AttendanceRecord.emp_id.in_(emp_ids))
    if from_date:
        q = q.filter(AttendanceRecord.date >= from_date)
    if to_date:
        q = q.filter(AttendanceRecord.date <= to_date)
    all_recs = q.all()

    # Build attendance dict adapter for stats computation
    attendance = {}
    for r in all_recs:
        if r.check_in_time:
            attendance.setdefault(r.date, {})[r.emp_id] = {
                "check_in": r.check_in_time,
                "check_out": r.check_out_time,
            }

    dates = sorted(attendance.keys())

    emp_stats = {
        eid: {"name": e["name"], "role": e["role"], "days": 0, "minutes": 0, "late_days": 0}
        for eid, e in employees.items()
    }
    daily_counts = []
    for d in dates:
        day_data = attendance[d]
        present = 0
        for eid, rec in day_data.items():
            check_in = rec.get("check_in")
            check_out = rec.get("check_out")
            if not check_in:
                continue
            present += 1
            if eid in emp_stats:
                emp_stats[eid]["days"] += 1
                start = employees[eid]["schedule"].get("start", "09:00")
                if check_in > _time_threshold(start, 15):
                    emp_stats[eid]["late_days"] += 1
                if check_out:
                    ci = datetime.strptime(check_in, "%H:%M:%S")
                    co = datetime.strptime(check_out, "%H:%M:%S")
                    emp_stats[eid]["minutes"] += int((co - ci).total_seconds()) // 60
        daily_counts.append({"date": d, "count": present})

    for es in emp_stats.values():
        h = es["minutes"] // 60
        m = es["minutes"] % 60
        es["hours_str"] = f"{h}ч {m}мин"
        es["hours"] = round(es["minutes"] / 60, 1)

    return jsonify({
        "dates": dates,
        "daily_counts": daily_counts,
        "employee_stats": list(emp_stats.values()),
        "total_days": len(dates)
    })

# ─── Audit log page + API ─────────────────────────────────────────────────────

@app.route("/audit")
@require_role("superadmin")
def audit_page():
    """Superadmin-only audit log viewer page."""
    user = User.query.get(session.get("user_id"))
    username = user.username if user else ""
    return render_template("audit.html", username=username)


@app.route("/api/audit", methods=["GET"])
@require_role("superadmin")
def audit_api():
    """Return most recent 500 AuditLog rows; supports optional ?action= and ?actor= filters."""
    action_filter = request.args.get("action", "").strip()
    actor_filter = request.args.get("actor", "").strip()

    q = AuditLog.query.order_by(AuditLog.id.desc())
    if action_filter:
        q = q.filter(AuditLog.action.contains(action_filter))
    if actor_filter:
        q = q.filter(AuditLog.actor_username.contains(actor_filter))
    rows = q.limit(500).all()

    return jsonify([
        {
            "ts": r.ts,
            "actor_username": r.actor_username,
            "action": r.action,
            "target_type": r.target_type,
            "target_id": r.target_id,
            "old_value": r.old_value,
            "new_value": r.new_value,
        }
        for r in rows
    ])


# ─── Startup ──────────────────────────────────────────────────────────────────
# db.create_all() is idempotent — safe to call on every startup (D-16, Pitfall 2).
# Must run BEFORE init_config/init_users so tables exist (T-06-06).

with app.app_context():
    db.create_all()
    # Idempotent column migration: add emp_id to existing user tables (D-10).
    # create_all() does NOT alter existing tables (Pitfall 4); this guard handles upgrades.
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE user ADD COLUMN emp_id TEXT"))
            conn.commit()
    except sa_exc.OperationalError:
        pass  # Column already exists — safe to ignore
    # Idempotent column migration: add display_name to existing user tables.
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE user ADD COLUMN display_name TEXT"))
            conn.commit()
    except sa_exc.OperationalError:
        pass  # Column already exists — safe to ignore
    # PERF-01: Replace two separate column indexes with a single composite index.
    # Drop SQLAlchemy-default-named indexes; create composite (emp_id, date) index.
    # Uses IF EXISTS / IF NOT EXISTS for full idempotency on repeated startups.
    try:
        with db.engine.connect() as conn:
            conn.execute(text("DROP INDEX IF EXISTS ix_attendance_record_emp_id"))
            conn.execute(text("DROP INDEX IF EXISTS ix_attendance_record_date"))
            conn.execute(text(
                "CREATE INDEX IF NOT EXISTS ix_attendance_emp_date "
                "ON attendance_record (emp_id, date)"
            ))
            conn.commit()
    except sa_exc.OperationalError:
        pass  # Table may not exist yet on a fresh DB — create_all() will apply it
    init_config()
    init_users()

if __name__ == "__main__":
    train_recognizer()
    app.run(debug=True, host="0.0.0.0", port=5050)
