"""
Tests for SADM-01 / D-07: global multi-org T-13 Excel export.

Coverage:
  (a) GET /api/superadmin/export/xlsx?month=YYYY-MM returns 200 with xlsx content-type
      and a non-empty body; openpyxl can load it; workbook contains one sheet per
      seeded org with sheet titles matching org names (truncated to 31 chars).
  (b) When no organizations are seeded, GET returns a workbook containing a single
      sheet named 'Нет данных' (empty-workbook guard, Pitfall 8).
  (c) GET as org_admin returns 403 (T-10-18 — access control).
"""
import pytest
from io import BytesIO
from datetime import datetime
import openpyxl

from tests.conftest import (
    BCRYPT_HASH_SUPERADMIN,
    seed_users,
    seed_orgs,
    seed_depts,
    seed_employees,
    seed_attendance,
)

# ─── Constants ────────────────────────────────────────────────────────────────

SA_ID  = "uid-sa-sadm01"
OA_ID  = "uid-oa-sadm01"
ORG_A  = "org-sadm01-A"
ORG_B  = "org-sadm01-B"
DEPT_A = "dept-sadm01-A"
DEPT_B = "dept-sadm01-B"
EMP_1  = "emp-sadm01-1"
EMP_2  = "emp-sadm01-2"

# Use a fixed past month so test behaviour is independent of the system clock
TEST_MONTH = "2026-01-01"
MONTH_PARAM = "2026-01"


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _inject_superadmin(client):
    with client.session_transaction() as sess:
        sess["user_id"] = SA_ID
        sess["role"] = "superadmin"
        sess["org_id"] = None
        sess["dept_id"] = None


def _inject_org_admin(client):
    with client.session_transaction() as sess:
        sess["user_id"] = OA_ID
        sess["role"] = "org_admin"
        sess["org_id"] = ORG_A
        sess["dept_id"] = None


def _seed_base_users(tmp_data):
    """Seed one superadmin and one org_admin."""
    seed_users(tmp_data, {
        SA_ID: {
            "id": SA_ID,
            "username": "sa_sadm01",
            "password_hash": BCRYPT_HASH_SUPERADMIN,
            "role": "superadmin",
            "active": True,
            "org_id": None,
            "dept_id": None,
        },
        OA_ID: {
            "id": OA_ID,
            "username": "oa_sadm01",
            "password_hash": BCRYPT_HASH_SUPERADMIN,
            "role": "org_admin",
            "active": True,
            "org_id": ORG_A,
            "dept_id": None,
        },
    })


def _seed_two_orgs(tmp_data):
    """Seed two organizations, two departments, and two employees."""
    seed_orgs(tmp_data, {
        ORG_A: {"id": ORG_A, "name": "Организация Альфа", "description": ""},
        ORG_B: {"id": ORG_B, "name": "Организация Бета",  "description": ""},
    })
    seed_depts(tmp_data, {
        DEPT_A: {"id": DEPT_A, "org_id": ORG_A, "name": "Отдел А", "head_name": ""},
        DEPT_B: {"id": DEPT_B, "org_id": ORG_B, "name": "Отдел Б", "head_name": ""},
    })
    seed_employees(tmp_data, {
        EMP_1: {
            "id": EMP_1, "name": "Иванов Иван",
            "role": "employee", "label": 1, "face_count": 5,
            "registered_at": "2026-01-01T00:00:00",
            "org_id": ORG_A, "dept_id": DEPT_A,
        },
        EMP_2: {
            "id": EMP_2, "name": "Петров Пётр",
            "role": "employee", "label": 2, "face_count": 5,
            "registered_at": "2026-01-01T00:00:00",
            "org_id": ORG_B, "dept_id": DEPT_B,
        },
    })


# ─── Test (a): two orgs → one sheet per org ───────────────────────────────────

def test_export_xlsx_one_sheet_per_org(client, tmp_data):
    """GET returns 200 xlsx; workbook has one sheet per seeded org."""
    _seed_base_users(tmp_data)
    _seed_two_orgs(tmp_data)
    seed_attendance(tmp_data, [
        {
            "emp_id": EMP_1,
            "date": TEST_MONTH,
            "check_in_time": "09:00:00",
            "check_out_time": "18:00:00",
            "event_type": "check_in",
        }
    ])
    _inject_superadmin(client)

    rv = client.get(f"/api/superadmin/export/xlsx?month={MONTH_PARAM}")
    assert rv.status_code == 200, f"Expected 200, got {rv.status_code}: {rv.data[:200]}"
    # Content-type must be xlsx (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)
    assert "spreadsheetml" in rv.content_type or "officedocument" in rv.content_type or rv.content_type == "application/octet-stream", \
        f"Unexpected content-type: {rv.content_type}"
    assert len(rv.data) > 0, "Response body is empty"

    wb = openpyxl.load_workbook(BytesIO(rv.data))
    assert len(wb.sheetnames) == 2, f"Expected 2 sheets, got {wb.sheetnames}"

    # Sheet titles must correspond to org names (truncated to 31 chars)
    expected_titles = {"Организация Альфа"[:31], "Организация Бета"[:31]}
    assert set(wb.sheetnames) == expected_titles, \
        f"Sheet names {wb.sheetnames} don't match orgs {expected_titles}"


# ─── Test (b): no orgs → placeholder sheet ────────────────────────────────────

def test_export_xlsx_no_orgs_returns_placeholder(client, tmp_data):
    """When no organizations are seeded, workbook contains exactly the 'Нет данных' sheet."""
    _seed_base_users(tmp_data)
    # Deliberately seed NO orgs, depts, or employees
    _inject_superadmin(client)

    rv = client.get(f"/api/superadmin/export/xlsx?month={MONTH_PARAM}")
    assert rv.status_code == 200, f"Expected 200, got {rv.status_code}"
    assert len(rv.data) > 0, "Response body is empty"

    wb = openpyxl.load_workbook(BytesIO(rv.data))
    assert len(wb.sheetnames) == 1, f"Expected exactly 1 sheet, got {wb.sheetnames}"
    assert wb.sheetnames[0] == "Нет данных", \
        f"Expected placeholder sheet 'Нет данных', got {wb.sheetnames[0]!r}"


# ─── Test (c): org_admin gets 403 ─────────────────────────────────────────────

def test_export_xlsx_403_for_org_admin(client, tmp_data):
    """GET /api/superadmin/export/xlsx as org_admin returns 403 (T-10-18)."""
    _seed_base_users(tmp_data)
    _inject_org_admin(client)

    rv = client.get(f"/api/superadmin/export/xlsx?month={MONTH_PARAM}")
    assert rv.status_code == 403, f"Expected 403, got {rv.status_code}"
